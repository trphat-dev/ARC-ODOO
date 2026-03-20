# -*- coding: utf-8 -*-
import json
import io
from datetime import datetime
from odoo import http
from odoo.http import request, content_disposition
from odoo.addons.user_permission_management.utils.permission_checker import require_module_access

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    openpyxl = None


class InvestorReportController(http.Controller):

    @http.route('/investor_report', type='http', auth='user', website=True)
    @require_module_access('report_list')
    def investor_report_page(self, **kws):
        """Render the main report page for Investor Report."""
        return request.render('report_list.investor_report_page_template', {})

    @http.route('/investor_report/get_data', type='json', auth='user', methods=['POST'])
    def get_investor_report_data(self, filters, page, limit, search_values=None, **kws):
        """API endpoint to fetch paginated report data based on filters."""
        try:
            # Lấy dữ liệu từ investor.list model
            domain = []
            
            # Filter theo ngày
            if filters.get('dateFrom'):
                domain.append(('open_date', '>=', filters['dateFrom']))
            if filters.get('dateTo'):
                domain.append(('open_date', '<=', filters['dateTo']))
            
            # Filter theo trạng thái
            if filters.get('status'):
                if filters['status'] == 'active':
                    domain.append(('status', 'in', ['kyc', 'vsd']))
                elif filters['status'] == 'inactive':
                    domain.append(('status', 'in', ['pending', 'incomplete']))
            
            # Search values
            if search_values:
                if search_values.get('account_number'):
                    domain.append(('account_number', 'ilike', search_values['account_number']))
                if search_values.get('customer_name'):
                    domain.append(('partner_name', 'ilike', search_values['customer_name']))

            
            # Lấy records với pagination
            offset = (int(page) - 1) * int(limit)
            records = request.env['investor.list'].search(domain, limit=int(limit), offset=offset)
            total_records = request.env['investor.list'].search_count(domain)
            
            # Chuẩn hóa dữ liệu
            data = []
            for i, record in enumerate(records):
                # Lấy thông tin từ partner
                partner = record.partner_id
                
                # Lấy thông tin từ investor.profile nếu có
                investor_profile = None
                if partner:
                    investor_profile = request.env['investor.profile'].search([
                        ('partner_id', '=', partner.id)
                    ], limit=1)
                
                # Lấy thông tin từ status.info
                status_info = None
                if partner:
                    status_info = request.env['status.info'].search([
                        ('partner_id', '=', partner.id)
                    ], limit=1)
                
                # Lấy thông tin địa chỉ
                permanent_address = ''
                contact_address = ''
                if partner:
                    address = request.env['investor.address'].search([
                        ('partner_id', '=', partner.id),
                        ('address_type', '=', 'permanent')
                    ], limit=1)
                    if address:
                        permanent_address = f"{address.street}, {address.ward}, {address.district}, {address.state_id.name if address.state_id else ''}"
                    
                    # Lấy địa chỉ liên hệ
                    contact_address_obj = request.env['investor.address'].search([
                        ('partner_id', '=', partner.id),
                        ('address_type', '=', 'current')
                    ], limit=1)
                    if contact_address_obj:
                        contact_address = f"{contact_address_obj.street}, {contact_address_obj.ward}, {contact_address_obj.district}, {contact_address_obj.state_id.name if contact_address_obj.state_id else ''}"
                
                # Xác định trạng thái tài khoản
                account_status = 'active' if record.status in ['kyc', 'vsd'] else 'inactive'
                
                # Lấy ngày đóng tài khoản (nếu có)
                close_date = ''
                if account_status == 'inactive' and record.write_date:
                    close_date = record.write_date.strftime('%d/%m/%Y')
                

                

                
                data.append({
                    'id': record.id,
                    'stt': offset + i + 1,
                    'account_number': record.account_number or '',
                    'trading_account': '',  # Không có trong investor.list
                    'customer_name': record.partner_name or '',
                    'id_number': record.id_number or '',
                    'id_issue_date': investor_profile.id_issue_date.strftime('%d/%m/%Y') if investor_profile and investor_profile.id_issue_date else '',
                    'id_issue_place': investor_profile.id_issue_place if investor_profile else '',
                    'permanent_address': permanent_address,
                    'contact_address': contact_address,
                    'phone_number': record.phone or '',
                    'email': record.email or '',
                    'open_date': record.open_date.strftime('%d/%m/%Y') if record.open_date else '',
                    'close_date': close_date,
                    'status': account_status,

                })
            
            return {
                'data': data,
                'total': total_records,
                'page': int(page),
                'limit': int(limit)
            }
            
        except Exception as e:
            # print(f"Error in get_investor_report_data: {e}")
            return {
                'data': [],
                'total': 0,
                'page': int(page),
                'limit': int(limit),
                'error': 'Internal server error'
            }
    
    @http.route('/investor_report/export_pdf', type='http', auth='user', website=True, csrf=False)
    def export_pdf(self, **kw):
        """Export report data as a PDF file."""
        try:
            # Lấy dữ liệu thật từ investor.list
            domain = []
            
            # Filter theo ngày
            if kw.get('dateFrom'):
                domain.append(('open_date', '>=', kw['dateFrom']))
            if kw.get('dateTo'):
                domain.append(('open_date', '<=', kw['dateTo']))
            
            # Filter theo trạng thái
            if kw.get('status'):
                if kw['status'] == 'active':
                    domain.append(('status', 'in', ['kyc', 'vsd']))
                elif kw['status'] == 'inactive':
                    domain.append(('status', 'in', ['pending', 'incomplete']))
            
            # Search values
            if kw.get('account_number'):
                domain.append(('account_number', 'ilike', kw['account_number']))
            if kw.get('customer_name'):
                domain.append(('partner_name', 'ilike', kw['customer_name']))

            
            records = request.env['investor.list'].search(domain)
            
            # Chuẩn hóa dữ liệu cho export
            export_data = []
            for i, record in enumerate(records):
                # Lấy thông tin từ partner
                partner = record.partner_id
                
                # Lấy thông tin từ investor.profile nếu có
                investor_profile = None
                if partner:
                    investor_profile = request.env['investor.profile'].search([
                        ('partner_id', '=', partner.id)
                    ], limit=1)
                
                # Lấy thông tin từ status.info
                status_info = None
                if partner:
                    status_info = request.env['status.info'].search([
                        ('partner_id', '=', partner.id)
                    ], limit=1)
                
                # Lấy thông tin địa chỉ
                permanent_address = ''
                contact_address = ''
                if partner:
                    address = request.env['investor.address'].search([
                        ('partner_id', '=', partner.id),
                        ('address_type', '=', 'permanent')
                    ], limit=1)
                    if address:
                        permanent_address = f"{address.street}, {address.ward}, {address.district}, {address.state_id.name if address.state_id else ''}"
                    
                    # Lấy địa chỉ liên hệ
                    contact_address_obj = request.env['investor.address'].search([
                        ('partner_id', '=', partner.id),
                        ('address_type', '=', 'current')
                    ], limit=1)
                    if contact_address_obj:
                        contact_address = f"{contact_address_obj.street}, {contact_address_obj.ward}, {contact_address_obj.district}, {contact_address_obj.state_id.name if contact_address_obj.state_id else ''}"
                
                # Xác định trạng thái tài khoản
                account_status = 'active' if record.status in ['kyc', 'vsd'] else 'inactive'
                
                # Lấy ngày đóng tài khoản (nếu có)
                close_date = ''
                if account_status == 'inactive' and record.write_date:
                    close_date = record.write_date.strftime('%d/%m/%Y')
                

                

                
                export_data.append({
                    'stt': i + 1,
                    'account_number': record.account_number or '',
                    'trading_account': '',
                    'customer_name': record.partner_name or '',
                    'id_number': record.id_number or '',
                    'id_issue_date': investor_profile.id_issue_date.strftime('%d/%m/%Y') if investor_profile and investor_profile.id_issue_date else '',
                    'id_issue_place': investor_profile.id_issue_place if investor_profile else '',
                    'permanent_address': permanent_address,
                    'contact_address': contact_address,
                    'phone_number': record.phone or '',
                    'email': record.email or '',
                    'open_date': record.open_date.strftime('%d/%m/%Y') if record.open_date else '',
                    'close_date': close_date,
                    'status': account_status,

                })
            
            records = export_data
            
        except Exception as e:
            # print(f"Error in export_pdf: {e}")
            records = []
        
        report_date_range = "Tất cả thời gian"
        date_from = kw.get('date_from')
        date_to = kw.get('date_to')
        if date_from and date_to:
            date_from_str = datetime.strptime(date_from, '%Y-%m-%d').strftime('%d/%m/%Y')
            date_to_str = datetime.strptime(date_to, '%Y-%m-%d').strftime('%d/%m/%Y')
            report_date_range = f"Từ ngày {date_from_str} đến ngày {date_to_str}"
        elif date_from:
            report_date_range = f"Từ ngày {datetime.strptime(date_from, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        elif date_to:
            report_date_range = f"Đến ngày {datetime.strptime(date_to, '%Y-%m-%d').strftime('%d/%m/%Y')}"

        pdf_data = {
            'records': records,
            'filters': {
                'date_from': date_from,
                'date_to': date_to,

                'search_term': kw.get('search_term')
            },
            'report_date_range': report_date_range,
            'report_date': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            'company': request.env.company,
        }
        
        pdf_content, content_type = request.env['ir.actions.report']._render_qweb_pdf(
            'report_list.investor_report_pdf_template',
            res_ids=[],
            data=pdf_data
        )
        
        filename = f"DanhSach_NhaDauTu_{datetime.now().strftime('%Y%m%d')}.pdf"
        pdf_http_headers = [
            ('Content-Type', 'application/pdf'),
            ('Content-Length', len(pdf_content)),
            ('Content-Disposition', content_disposition(filename)),
        ]
        return request.make_response(pdf_content, headers=pdf_http_headers)

    @http.route('/investor_report/export_xlsx', type='http', auth='user', website=True, csrf=False)
    def export_xlsx(self, **kw):
        """Export report data as an XLSX file."""
        if not openpyxl:
            return request.make_response(
                "Thư viện 'openpyxl' chưa được cài đặt. Vui lòng cài đặt bằng lệnh 'pip install openpyxl'.",
                headers=[('Content-Type', 'text/plain')]
            )

        try:
            # Lấy dữ liệu thật từ investor.list
            domain = []
            
            # Filter theo ngày
            if kw.get('dateFrom'):
                domain.append(('open_date', '>=', kw['dateFrom']))
            if kw.get('dateTo'):
                domain.append(('open_date', '<=', kw['dateTo']))
            
            # Filter theo trạng thái
            if kw.get('status'):
                if kw['status'] == 'active':
                    domain.append(('status', 'in', ['kyc', 'vsd']))
                elif kw['status'] == 'inactive':
                    domain.append(('status', 'in', ['pending', 'incomplete']))
            
            # Search values
            if kw.get('account_number'):
                domain.append(('account_number', 'ilike', kw['account_number']))
            if kw.get('customer_name'):
                domain.append(('partner_name', 'ilike', kw['customer_name']))

            
            records = request.env['investor.list'].search(domain)
            
            # Chuẩn hóa dữ liệu cho export
            export_data = []
            for i, record in enumerate(records):
                # Lấy thông tin từ partner
                partner = record.partner_id
                
                # Lấy thông tin từ investor.profile nếu có
                investor_profile = None
                if partner:
                    investor_profile = request.env['investor.profile'].search([
                        ('partner_id', '=', partner.id)
                    ], limit=1)
                
                # Lấy thông tin từ status.info
                status_info = None
                if partner:
                    status_info = request.env['status.info'].search([
                        ('partner_id', '=', partner.id)
                    ], limit=1)
                
                # Lấy thông tin địa chỉ
                permanent_address = ''
                contact_address = ''
                if partner:
                    address = request.env['investor.address'].search([
                        ('partner_id', '=', partner.id),
                        ('address_type', '=', 'permanent')
                    ], limit=1)
                    if address:
                        permanent_address = f"{address.street}, {address.ward}, {address.district}, {address.state_id.name if address.state_id else ''}"
                    
                    # Lấy địa chỉ liên hệ
                    contact_address_obj = request.env['investor.address'].search([
                        ('partner_id', '=', partner.id),
                        ('address_type', '=', 'current')
                    ], limit=1)
                    if contact_address_obj:
                        contact_address = f"{contact_address_obj.street}, {contact_address_obj.ward}, {contact_address_obj.district}, {contact_address_obj.state_id.name if contact_address_obj.state_id else ''}"
                
                # Xác định trạng thái tài khoản
                account_status = 'active' if record.status in ['kyc', 'vsd'] else 'inactive'
                
                # Lấy ngày đóng tài khoản (nếu có)
                close_date = ''
                if account_status == 'inactive' and record.write_date:
                    close_date = record.write_date.strftime('%d/%m/%Y')
                

                

                
                export_data.append({
                    'stt': i + 1,
                    'account_number': record.account_number or '',
                    'trading_account': '',
                    'customer_name': record.partner_name or '',
                    'id_number': record.id_number or '',
                    'id_issue_date': investor_profile.id_issue_date.strftime('%d/%m/%Y') if investor_profile and investor_profile.id_issue_date else '',
                    'id_issue_place': investor_profile.id_issue_place if investor_profile else '',
                    'permanent_address': permanent_address,
                    'contact_address': contact_address,
                    'phone_number': record.phone or '',
                    'email': record.email or '',
                    'open_date': record.open_date.strftime('%d/%m/%Y') if record.open_date else '',
                    'close_date': close_date,
                    'status': account_status,

                })
            
            records = export_data
            
        except Exception as e:
            # print(f"Error in export_xlsx: {e}")
            records = []

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "DanhSachNhaDauTu"

        # Header
        headers = [
            'STT', 'Số TK', 'TK GDCK', 'Khách hàng', 'Số CCCD/CC/GPKD', 'Ngày cấp', 'Nơi cấp', 
            'Địa chỉ thường trú', 'Địa chỉ liên hệ', 'Số điện thoại', 'Email', 
            'Trạng thái'
        ]
        sheet.append(headers)

        # Style Header
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Body
        for rec in records:
            row = [
                rec.get('stt'), rec.get('account_number'), rec.get('trading_account'), rec.get('customer_name'),
                rec.get('id_number'), rec.get('id_issue_date'), rec.get('id_issue_place'), rec.get('permanent_address'),
                rec.get('contact_address'), rec.get('phone_number'), rec.get('email'), rec.get('status')
            ]
            sheet.append(row)

        # Auto-fit column width
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            sheet.column_dimensions[column].width = adjusted_width

        # Save to buffer
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f"DanhSach_NhaDauTu_{datetime.now().strftime('%Y%m%d')}.xlsx"
        xlsx_headers = [
            ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Disposition', content_disposition(filename)),
        ]
        return request.make_response(output.read(), headers=xlsx_headers)
