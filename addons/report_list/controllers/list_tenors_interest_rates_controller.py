# -*- coding: utf-8 -*-
import io
from datetime import datetime
from odoo import http
from odoo.http import request, content_disposition
from odoo.addons.user_permission_management.utils.permission_checker import require_module_access

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    # *** FIX: Import get_column_letter utility ***
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class ListTenorsInterestRatesController(http.Controller):

    @http.route('/list_tenors_interest_rates', type='http', auth='user', website=True)
    @require_module_access('report_list')
    def report_page(self, **kws):
        """Hiển thị trang báo cáo chính chứa component Owl."""
        return request.render('report_list.list_tenors_interest_rates_page_template', {})

    @http.route('/list_tenors_interest_rates/get_data', type='json', auth='user', methods=['POST'])
    def get_report_data(self, filters, page, limit, search_values=None, **kws):
        """API endpoint để lấy dữ liệu báo cáo đã phân trang từ nav.term.rate."""
        try:
            # Lấy dữ liệu từ nav.term.rate model
            domain = []
            
            # Filter theo kỳ hạn
            if filters.get('term'):
                domain.append(('term_months', '=', int(filters['term'])))
            
            # Filter theo ngày hiệu lực
            if filters.get('dateFrom'):
                domain.append(('effective_date', '>=', filters['dateFrom']))
            if filters.get('dateTo'):
                domain.append(('effective_date', '<=', filters['dateTo']))
            
            # Search values
            if search_values:
                if search_values.get('product_name'):
                    domain.append(('description', 'ilike', search_values['product_name']))
            
            # Lấy records với pagination
            offset = (int(page) - 1) * int(limit)
            records = request.env['nav.term.rate'].search(domain, limit=int(limit), offset=offset)
            total_records = request.env['nav.term.rate'].search_count(domain)
            
            # Chuẩn hóa dữ liệu
            data = []
            for i, record in enumerate(records):
                data.append({
                    'id': record.id,
                    'stt': offset + i + 1,
                    'product_name': record.description or f'Kỳ hạn {record.term_months} tháng',
                    'product_creation_date': record.create_date.strftime('%d/%m/%Y') if record.create_date else '',
                    'effective_date': record.effective_date.strftime('%d/%m/%Y') if record.effective_date else '',
                    'end_date': record.end_date.strftime('%d/%m/%Y') if record.end_date else '',
                    'term': f'{record.term_months} tháng',
                    'interest_rate': f'{record.interest_rate}%',
                    'creator_name': record.create_uid.name if record.create_uid else '',
                    'status': 'active' if record.active else 'inactive'
                })
            
            return {
                'data': data,
                'total': total_records,
                'page': int(page),
                'limit': int(limit)
            }
            
        except Exception as e:
            print(f"Error in get_report_data: {e}")
            return {
                'data': [],
                'total': 0,
                'page': int(page),
                'limit': int(limit),
                'error': str(e)
            }
        
    @http.route('/list_tenors_interest_rates/get_filter_options', type='json', auth='user', methods=['POST'])
    def get_filter_options(self, **kws):
        """API endpoint để lấy các tùy chọn cho bộ lọc dropdown từ nav.term.rate."""
        try:
            # Tạo danh sách kỳ hạn đầy đủ từ 1-12 tháng
            term_options = []
            for i in range(1, 13):
                term_options.append({
                    'value': i,
                    'label': f'{i} tháng'
                })
            
            return {
                'term_options': term_options
            }
            
        except Exception as e:
            print(f"Error in get_filter_options: {e}")
            return {
                'term_options': []
            }

    def _get_filters_from_kw(self, kw):
        """Helper để trích xuất filters từ keyword arguments."""
        return {
            'dateFrom': kw.get('dateFrom'),
            'dateTo': kw.get('dateTo'),
            'term': kw.get('term'),
            'product_name': kw.get('product_name')
        }

    def _get_report_date_range(self, filters):
        """Tạo chuỗi mô tả khoảng thời gian báo cáo."""
        date_from = filters.get('dateFrom')
        date_to = filters.get('dateTo')

        if date_from or date_to:
            date_from_str = datetime.strptime(date_from, '%Y-%m-%d').strftime('%d/%m/%Y') if date_from else '...'
            date_to_str = datetime.strptime(date_to, '%Y-%m-%d').strftime('%d/%m/%Y') if date_to else '...'
            return f"Khoảng thời gian từ {date_from_str} đến {date_to_str}"
            
        return "Toàn bộ thời gian"


    @http.route('/list_tenors_interest_rates/export_pdf', type='http', auth='user', website=True, csrf=False)
    def export_pdf(self, **kw):
        """Xuất dữ liệu báo cáo ra file PDF từ nav.term.rate."""
        try:
            # Lấy dữ liệu từ nav.term.rate
            domain = []
            
            # Filter theo kỳ hạn
            if kw.get('term'):
                domain.append(('term_months', '=', int(kw['term'])))
            
            # Filter theo ngày hiệu lực
            if kw.get('dateFrom'):
                domain.append(('effective_date', '>=', kw['dateFrom']))
            if kw.get('dateTo'):
                domain.append(('effective_date', '<=', kw['dateTo']))
            
            records = request.env['nav.term.rate'].search(domain)
            
            # Chuẩn hóa dữ liệu cho export
            export_data = []
            for i, record in enumerate(records):
                export_data.append({
                    'stt': i + 1,
                    'product_name': record.description or f'Kỳ hạn {record.term_months} tháng',
                    'product_creation_date': record.create_date.strftime('%d/%m/%Y') if record.create_date else '',
                    'effective_date': record.effective_date.strftime('%d/%m/%Y') if record.effective_date else '',
                    'end_date': record.end_date.strftime('%d/%m/%Y') if record.end_date else '',
                    'term': f'{record.term_months} tháng',
                    'interest_rate': f'{record.interest_rate}%',
                    'creator_name': record.create_uid.name if record.create_uid else ''
                })
            
            report_date_range = self._get_report_date_range(kw)

            pdf_data = {
                'records': export_data,
                'report_date_range': report_date_range,
                'report_date': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            }
            
            pdf_content, content_type = request.env['ir.actions.report']._render_qweb_pdf(
                'report_list.list_tenors_interest_rates_pdf_template',
                res_ids=[],
                data=pdf_data
            )
            
            filename = f"DS_KyHan_LaiSuat_{datetime.now().strftime('%Y%m%d')}.pdf"
            pdf_http_headers = [
                ('Content-Type', 'application/pdf'),
                ('Content-Length', len(pdf_content)),
                ('Content-Disposition', content_disposition(filename)),
            ]
            return request.make_response(pdf_content, headers=pdf_http_headers)
            
        except Exception as e:
            print(f"Error in export_pdf: {e}")
            return request.make_response(
                f"Lỗi khi xuất PDF: {str(e)}",
                headers=[('Content-Type', 'text/plain')]
            )

    @http.route('/list_tenors_interest_rates/export_xlsx', type='http', auth='user', website=True, csrf=False)
    def export_xlsx(self, **kw):
        """Xuất dữ liệu báo cáo ra file XLSX từ nav.term.rate."""
        if not openpyxl:
            return request.make_response(
                "Thư viện 'openpyxl' chưa được cài đặt. Vui lòng cài đặt bằng lệnh 'pip install openpyxl'.",
                headers=[('Content-Type', 'text/plain')]
            )

        try:
            # Lấy dữ liệu từ nav.term.rate
            domain = []
            
            # Filter theo kỳ hạn
            if kw.get('term'):
                domain.append(('term_months', '=', int(kw['term'])))
            
            # Filter theo ngày hiệu lực
            if kw.get('dateFrom'):
                domain.append(('effective_date', '>=', kw['dateFrom']))
            if kw.get('dateTo'):
                domain.append(('effective_date', '<=', kw['dateTo']))
            
            records = request.env['nav.term.rate'].search(domain)

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "DS_KyHan_LaiSuat"

            # Tiêu đề báo cáo
            sheet.merge_cells('A1:H1')
            title_cell = sheet['A1']
            title_cell.value = "DANH SÁCH KỲ HẠN VÀ LÃI SUẤT"
            title_cell.font = Font(name='Arial', size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Phụ đề (khoảng thời gian)
            sheet.merge_cells('A2:H2')
            subtitle_cell = sheet['A2']
            subtitle_cell.value = self._get_report_date_range(kw)
            subtitle_cell.font = Font(name='Arial', size=10, italic=True)
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.row_dimensions[2].height = 20
            
            # Header bảng (bắt đầu từ dòng 4, sau 1 dòng trống)
            headers = [
                'STT', 'Sản phẩm (SP)', 'Ngày tạo SP', 'Ngày hiệu lực', 
                'Ngày kết thúc', 'Kỳ hạn', 'Lãi suất (%)', 'Người nhập/Điều chỉnh'
            ]
            sheet.append(headers) 

            # Style Header
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='f37731', end_color='f37731', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in sheet[4]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Body - Chuẩn hóa dữ liệu từ nav.term.rate
            for i, record in enumerate(records):
                row = [
                    i + 1,  # STT
                    record.description or f'Kỳ hạn {record.term_months} tháng',  # Sản phẩm
                    record.create_date.strftime('%d/%m/%Y') if record.create_date else '',  # Ngày tạo SP
                    record.effective_date.strftime('%d/%m/%Y') if record.effective_date else '',  # Ngày hiệu lực
                    record.end_date.strftime('%d/%m/%Y') if record.end_date else '',  # Ngày kết thúc
                    f'{record.term_months} tháng',  # Kỳ hạn
                    f'{record.interest_rate}%',  # Lãi suất
                    record.create_uid.name if record.create_uid else ''  # Người nhập/Điều chỉnh
                ]
                sheet.append(row)

            # *** FIX START: Use a safe loop for auto-fitting column width ***
            for col_idx in range(1, sheet.max_column + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in sheet[column_letter]:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = (max_length + 2) if max_length < 40 else 40
                sheet.column_dimensions[column_letter].width = adjusted_width
            # *** FIX END ***

            # Save to buffer
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            filename = f"DS_KyHan_LaiSuat_{datetime.now().strftime('%Y%m%d')}.xlsx"
            xlsx_headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition(filename)),
            ]
            return request.make_response(output.read(), headers=xlsx_headers)
            
        except Exception as e:
            print(f"Error in export_xlsx: {e}")
            return request.make_response(
                f"Lỗi khi xuất XLSX: {str(e)}",
                headers=[('Content-Type', 'text/plain')]
            )
