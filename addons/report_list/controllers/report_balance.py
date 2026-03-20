import json
import csv
import io
from datetime import datetime
from collections import defaultdict
from types import SimpleNamespace
from odoo import http, fields
from odoo.http import request
from odoo.addons.user_permission_management.utils.permission_checker import require_module_access

class ReportBalanceController(http.Controller):
    @http.route('/report-balance', type='http', auth='user', website=True)
    @require_module_access('report_list')
    def report_balance_page(self, **kw):
        """Render Report Balance page."""
        return request.render('report_list.report_balance_page_template', {})

    def _build_search_domain(self, domain, search_values, filters):
        """Normalize incoming domain/search params for portfolio.transaction."""
        normalized = list(domain or [])

        field_mapping = {
            'account_number': 'user_id.partner_id.name',
            'investor_name': 'user_id.name',
            'fund_name': 'fund_id.name',
        }

        if search_values:
            for field, value in search_values.items():
                if value:
                    actual_field = field_mapping.get(field, field)
                    normalized.append((actual_field, 'ilike', value))

        filters = filters or {}
        fund_filter = filters.get('fund')
        date_from = filters.get('dateFrom') or filters.get('from_date')
        date_to = filters.get('dateTo') or filters.get('to_date')

        if fund_filter:
            try:
                fund_id_int = int(fund_filter)
            except Exception:
                fund_id_int = fund_filter
            normalized.append(('fund_id', '=', fund_id_int))

        def _to_datetime(value, start=True):
            if not value:
                return False
            suffix = "00:00:00" if start else "23:59:59"
            return f"{value} {suffix}"

        if date_from and date_to and date_from == date_to:
            normalized.append(('create_date', '>=', _to_datetime(date_from, True)))
            normalized.append(('create_date', '<=', _to_datetime(date_to, False)))
        else:
            if date_from:
                normalized.append(('create_date', '>=', _to_datetime(date_from, True)))
            if date_to:
                normalized.append(('create_date', '<=', _to_datetime(date_to, False)))

        # Removed non-existent fields for portfolio.investment (transaction_type)
        return normalized



    @http.route('/report-balance/data', type='json', auth='user', methods=['POST'])
    def get_report_data(self, domain=None, search_values=None, limit=10, offset=0, **kw):
        """JSON RPC endpoint to fetch report balance data."""
        try:
            page = max(int(kw.get('page', 1) or 1), 1)
            limit = int(limit or 10)
            offset = int(kw.get('offset', (page - 1) * limit))

            filters = kw.get('filters') or {}
            
            # Domain for portfolio.investment
            normalized_domain = self._build_search_domain(domain, search_values, filters)
            
            normalized_domain.append(('status', '=', 'active'))
            
            Model = request.env['portfolio.investment']
            
            # Count total
            total = Model.search_count(normalized_domain)
            
            # Search records
            fields_to_read = [
                'id',
                'report_account_number',
                'report_investor_name', 
                'report_id_number',
                'report_email',
                'report_phone_number',
                'fund_id',          # returns (id, name)
                'units',            # report quantity
                'total_value',      # report amount 
            ]
            
            # Order by latest created or updated
            records = Model.search_read(
                normalized_domain, 
                fields_to_read, 
                offset=offset, 
                limit=limit, 
                order='create_date desc'
            )
            
            # Post-process to flatten data for frontend
            data = []
            for rec in records:
                fund = rec.get('fund_id') 
                fund_name = fund[1] if fund else ''
                
                data.append({
                    'id': rec['id'],
                    'account_number': rec['report_account_number'],
                    'trading_account': rec['report_account_number'], 
                    'investor_name': rec['report_investor_name'],
                    'id_number': rec['report_id_number'],
                    'phone_number': rec['report_phone_number'],
                    'email': rec['report_email'],
                    'program_ticker': fund_name, 
                    'fund_name': fund_name,
                    'ccq_quantity': rec['units'],
                    'amount': rec['total_value'], 
                })
            return {
                'data': data,
                'total': total,
                'page': page,
                'limit': limit
            }

        except Exception as e:
            # print(f"Error in get_report_data: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'error': 'Internal server error',
                'records': [],
                'total': 0,
            }

    @http.route('/report-balance/products', type='json', auth='user', methods=['POST'])
    def get_products(self, **kw):
        """Return list of funds for filters."""
        try:
            funds = request.env['portfolio.fund'].search_read([], ['id', 'name', 'ticker'], order='name')
            return [{'id': fund['id'], 'name': fund['name'], 'ticker': fund.get('ticker')} for fund in funds]
        except Exception as e:
            # print(f"Error in get_products: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    def _get_export_rows(self, filters):
        # We replace this to return recordset or list of dicts from recordset
        Model = request.env['portfolio.investment']
        domain = self._build_search_domain([], {}, filters)
        domain.append(('status', '=', 'active'))
        records = Model.search(domain, order='create_date desc')
        
        rows = []
        for rec in records:
            rows.append({
                'account_number': rec.report_account_number,
                'investor_name': rec.report_investor_name,
                'phone_number': rec.report_phone_number,
                'id_number': rec.report_id_number,
                'email': rec.report_email,
                'fund_name': rec.fund_id.name if rec.fund_id else '',
                'program_ticker': rec.report_program_ticker,
                'ccq_quantity': rec.units,
                'amount': rec.total_value,
                'status': rec.status
            })
        return rows

    @http.route('/report-balance/export-pdf', type='http', auth='user', website=True)
    def export_pdf(self, **kw):
        """Export report data to PDF."""
        try:
            filters = {
                'fund': kw.get('product'),
                'dateTo': kw.get('to_date'),
                'dateFrom': kw.get('from_date'),
            }
            rows = self._get_export_rows(filters)

            records = [
                SimpleNamespace(
                    account_number=row['account_number'],
                    investor_name=row['investor_name'],
                    id_number=row['id_number'],
                    fund_name=row['fund_name'],
                    fund_ticker=row['program_ticker'],
                    units=row['ccq_quantity'],
                    amount=row.get('amount', 0),
                    status=row.get('status', 'active')
                )
                for row in rows
            ]


            data = {
                'records': records,
                'total_records': len(records),
                'selected_product': filters.get('fund'),
                'to_date': filters.get('dateTo'),
                'export_time': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'company_name': request.env.company.name,
                'user_name': request.env.user.name,
            }

            pdf_content, content_type = request.env['ir.actions.report']._render_qweb_pdf(
                'report_list.report_balance_pdf_template',
                res_ids=[request.env.user.id],
                data=data
            )

            pdf_http_headers = [
                ('Content-Type', 'application/pdf'),
                ('Content-Length', len(pdf_content)),
                ('Content-Disposition', f'attachment; filename="Bao_cao_so_du_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'),
            ]
            return request.make_response(pdf_content, headers=pdf_http_headers)

        except Exception as e:
            # print(f"Error in export_pdf: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                "<h1>Error</h1><p>An internal error occurred. Please try again later.</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-balance/export-xlsx', type='http', auth='user', website=True)
    def export_xlsx(self, **kw):
        """Export report data to XLSX."""
        try:
            filters = {
                'fund': kw.get('product'),
                'dateFrom': kw.get('from_date'),
                'dateTo': kw.get('to_date'),
            }
            rows = self._get_export_rows(filters)

            import openpyxl
            from openpyxl.styles import Font, Alignment

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Report Balance"

            headers = ['STT', 'Số TK', 'Tên khách hàng', 'Số điện thoại', 'Email', 'Quỹ', 'Mã CK', 'Số CCQ']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            for idx, row in enumerate(rows, start=2):
                sheet.cell(row=idx, column=1, value=idx-1)
                sheet.cell(row=idx, column=2, value=row.get('account_number', ''))
                sheet.cell(row=idx, column=3, value=row.get('investor_name', ''))
                sheet.cell(row=idx, column=4, value=row.get('phone_number', ''))
                sheet.cell(row=idx, column=5, value=row.get('email', ''))
                sheet.cell(row=idx, column=6, value=row.get('fund_name', ''))
                sheet.cell(row=idx, column=7, value=row.get('program_ticker', ''))
                sheet.cell(row=idx, column=8, value=row.get('ccq_quantity', 0))

            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                column_letter = column_cells[0].column_letter
                sheet.column_dimensions[column_letter].width = max(10, length + 2)

            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            xlsx_headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="report_balance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'),
            ]
            return request.make_response(output.read(), headers=xlsx_headers)

        except Exception as e:
            # print(f"Error in export_xlsx: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                "<h1>Error</h1><p>An internal error occurred. Please try again later.</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-balance/export-csv', type='http', auth='user', website=True)
    def export_csv(self, **kw):
        """Export report data to CSV."""
        try:
            filters = {
                'fund': kw.get('product'),
                'dateFrom': kw.get('from_date'),
                'dateTo': kw.get('to_date'),
            }
            rows = self._get_export_rows(filters)

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Số tài khoản',
                'Nhà đầu tư',
                'Số điện thoại',
                'ĐKSH',
                'Email',
                'Quỹ',
                'Mã CK',
                'Số CCQ'
            ])

            for row in rows:
                writer.writerow([
                    row.get('account_number', ''),
                    row.get('investor_name', ''),
                    row.get('phone_number', ''),
                    row.get('id_number', ''),
                    row.get('email', ''),
                    row.get('fund_name', ''),
                    row.get('program_ticker', ''),
                    f"{row.get('ccq_quantity', 0):.2f}",
                ])

            csv_content = output.getvalue()
            output.close()

            csv_http_headers = [
                ('Content-Type', 'text/csv; charset=utf-8'),
                ('Content-Length', len(csv_content.encode('utf-8'))),
                ('Content-Disposition', f'attachment; filename="Bao_cao_so_du_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'),
            ]
            return request.make_response(csv_content, headers=csv_http_headers)

        except Exception as e:
            # print(f"Error in export_csv: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                "<h1>Error</h1><p>An internal error occurred. Please try again later.</p>",
                headers=[('Content-Type', 'text/html')]
            )

    def _format_currency(self, amount):
        """Helper method to format currency."""
        if not amount:
            return '0'
        return f"{amount:,.0f}".replace(',', '.')

    def _format_date(self, date_value):
        """Helper method to format date."""
        if not date_value:
            return ''
        if isinstance(date_value, str):
            try:
                date_value = datetime.strptime(date_value, '%Y-%m-%d').date()
            except Exception:
                return date_value
        return date_value.strftime('%d/%m/%Y')

    def _get_investor_type_label(self, investor_type):
        investor_types = {
            'truc_tiep': 'Trực tiếp',
            'ky_danh': 'Ký danh',
        }
        return investor_types.get(investor_type, investor_type)
