import json
import csv
import io
from datetime import datetime, timedelta
from odoo import http
from odoo.http import request
from odoo.addons.user_permission_management.utils.permission_checker import require_module_access

class ReportContractStatisticsController(http.Controller):
    @http.route('/report-contract-statistics', type='http', auth='user', website=True)
    @require_module_access('report_list')
    def report_contract_statistics_page(self, **kw):
        """
        Route to render the Report Contract Statistics page.
        """
        return request.render('report_list.report_contract_statistics_page_template', {})

    @http.route('/report-contract-statistics/data', type='json', auth='user', methods=['POST'])
    def get_report_data(self, domain=None, search_values=None, limit=10, offset=0, **kw):
        """
        JSON RPC endpoint to fetch contract statistics data from transaction_list.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Start with the base domain from filters
            if not domain:
                domain = []

            # Add search values from inline search fields
            if search_values:
                for field, value in search_values.items():
                    if value: # Only add to domain if there is a value
                        # Map frontend field names to actual model fields
                        field_mapping = {
                            'so_hop_dong': 'reference',
                            'so_tk': 'account_number',
                            'khach_hang': 'investor_name',
                            'nvcs': 'approved_by.name',
                        }
                        actual_field = field_mapping.get(field, field)
                        domain.append((actual_field, 'ilike', value))

            # Resolve pagination params (page -> offset)
            page = int(kw.get('page', 1) or 1)
            limit = int(limit or 10)
            offset = int(kw.get('offset', (page - 1) * limit))

            # Build domain from filters (fund, dateFrom, dateTo, term)
            filters = kw.get('filters') or {}
            fund_filter = filters.get('fund')
            date_from = filters.get('dateFrom') or filters.get('from_date')
            date_to = filters.get('dateTo') or filters.get('to_date')
            term = filters.get('term')

            if fund_filter:
                try:
                    fund_id_int = int(fund_filter)
                except Exception:
                    fund_id_int = fund_filter
                domain.append(('fund_id', '=', fund_id_int))
            if term:
                try:
                    domain.append(('term_months', '=', int(term)))
                except Exception:
                    pass
            if date_from and date_to and date_from == date_to:
                domain.append(('create_date', '>=', f"{date_from} 00:00:00"))
                domain.append(('create_date', '<=', f"{date_to} 23:59:59"))
            else:
                if date_from:
                    domain.append(('create_date', '>=', f"{date_from} 00:00:00"))
                if date_to:
                    domain.append(('create_date', '<=', f"{date_to} 23:59:59"))

            # Debug: Log domain and parameters
            print(f"Report Contract Statistics Domain: {domain}")
            print(f"Limit: {limit}, Offset: {offset}, Page: {page}, Filters: {filters}")
            
            # Get total count
            total = report_model.search_count(domain)
            
            # Get records with pagination
            records = report_model.search(domain, limit=limit, offset=offset, order='create_date desc')
            
            # Format data for frontend
            data = []
            for i, record in enumerate(records, 1):
                # Sử dụng các field NAV đã tính toán sẵn từ transaction record
                nav_maturity_date = getattr(record, 'nav_maturity_date', False)
                nav_sell_date = getattr(record, 'nav_sell_date', False)
                
                # Format ngày đến hạn: ưu tiên nav_maturity_date, fallback về tính toán cũ nếu không có
                ngay_den_han = ''
                if nav_maturity_date:
                    ngay_den_han = nav_maturity_date.strftime('%d/%m/%Y')
                elif record.create_date and record.term_months:
                    # Fallback: tính từ created_at + term_months (sử dụng relativedelta logic)
                    from dateutil.relativedelta import relativedelta
                    maturity_date = record.create_date.date() + relativedelta(months=record.term_months)
                    ngay_den_han = maturity_date.strftime('%d/%m/%Y')
                
                data.append({
                    'id': record.id,
                    'stt': i,
                    'so_hop_dong': record.reference or '',
                    'so_tk': record.account_number or '',
                    'so_tk_gdck': record.account_number or '',  # Same as so_tk
                    'khach_hang': record.investor_name or '',
                    'ky_han': record.term_months or 0,
                    'so_tien': record.amount or 0,
                    'ngay_hop_dong': record.create_date.strftime('%d/%m/%Y') if record.create_date else '',
                    'ngay_den_han': ngay_den_han,
                    'nvcs': record.approved_by.name if record.approved_by else '',
                    'don_vi': 'HDCapital',  # Default company
                })
            
            result = {
                'data': data,
                'total': total,
                'page': page,
                'limit': limit,
            }
            
            print(f"Returning data: {result}")
            return result
            
        except Exception as e:
            print(f"Error in get_report_data: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'error': str(e),
                'data': [],
                'total': 0,
            }

    @http.route('/report-contract-statistics/terms', type='json', auth='user', methods=['POST'])
    def get_terms(self, **kw):
        """
        JSON RPC endpoint to fetch terms for the filter dropdown.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Get distinct terms from transaction records
            terms = report_model.read_group([], ['term_months'], ['term_months'])
            products = []
            
            for term in terms:
                if term.get('term_months'):
                    term_value = term['term_months']
                    products.append({
                        'id': term_value, 
                        'name': f"{term_value} tháng"
                    })
            
            # Sort products by term
            products.sort(key=lambda x: x['id'])
            print(f"Terms found: {products}")
            return products
            
        except Exception as e:
            print(f"Error in get_terms: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    @http.route('/report-contract-statistics/export-pdf', type='http', auth='user', website=True)
    def export_pdf(self, **kw):
        """
        Route to export the contract statistics data as a PDF.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            selected_term = kw.get('term', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            don_vi = kw.get('don_vi', '')
            
            # Build domain based on filters
            domain = []
            if selected_term:
                domain.append(['term_months', '=', int(selected_term)])
            if from_date:
                domain.append(['create_date', '>=', from_date])
            if to_date:
                domain.append(['create_date', '<=', to_date])
            # don_vi filter not applicable for transaction data
            
            # Get all records for PDF export (no pagination)
            records = report_model.search(domain, order='create_date desc')
            
            # Prepare data for template
            data = {
                'records': records,
                'total_records': len(records),
                'selected_term': selected_term,
                'from_date': from_date,
                'to_date': to_date,
                'don_vi': don_vi,
                'export_time': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'company_name': request.env.company.name,
                'user_name': request.env.user.name,
            }
            
            # Generate PDF report
            pdf_content, content_type = request.env['ir.actions.report']._render_qweb_pdf(
                'report_list.report_contract_statistics_pdf_template',
                res_ids=[1],  # Pass a dummy ID
                data=data
            )
            
            # Set response headers
            pdf_http_headers = [
                ('Content-Type', 'application/pdf'),
                ('Content-Length', len(pdf_content)),
                ('Content-Disposition', f'attachment; filename="Bao_cao_thong_ke_HD_ky_han_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'),
            ]
            
            return request.make_response(pdf_content, headers=pdf_http_headers)
            
        except Exception as e:
            print(f"Error in export_pdf: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất PDF</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-contract-statistics/export-xlsx', type='http', auth='user', website=True)
    def export_xlsx(self, **kw):
        """
        Route to export the report contract statistics data as an XLSX.
        """
        try:
            # Sử dụng portfolio.transaction
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            customer = kw.get('customer', '')
            contract = kw.get('contract', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            
            # Build domain based on filters
            domain = []
            if customer:
                domain.append(['user_id.name', 'ilike', customer])
            if contract:
                domain.append(['name', 'ilike', contract])
            if from_date:
                domain.append(['created_at', '>=', f"{from_date} 00:00:00"])
            if to_date:
                domain.append(['created_at', '<=', f"{to_date} 23:59:59"])
            
            # Get all records for XLSX export (no pagination)
            records = report_model.search(domain, order='created_at desc')
            
            # Create XLSX content
            import io
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Report Contract Statistics"
            
            # Headers
            headers = ['STT', 'Số hợp đồng', 'Số TK', 'Khách hàng', 'Số tiền', 'Đơn vị']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for i, record in enumerate(records, 2):
                sheet.cell(row=i, column=1, value=i-1)  # STT
                sheet.cell(row=i, column=2, value=record.name or '')
                sheet.cell(row=i, column=3, value=record.user_id.name if record.user_id else '')
                sheet.cell(row=i, column=4, value=record.user_id.name if record.user_id else '')
                sheet.cell(row=i, column=5, value=record.amount or 0)
                sheet.cell(row=i, column=6, value='VND')
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to buffer
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            # Set response headers
            xlsx_headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="report_contract_statistics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'),
            ]
            
            return request.make_response(output.read(), headers=xlsx_headers)
            
        except Exception as e:
            print(f"Error in export_xlsx: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất XLSX</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-contract-statistics/export-csv', type='http', auth='user', website=True)
    def export_csv(self, **kw):
        """
        Route to export the contract statistics data as a CSV.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            selected_term = kw.get('term', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            don_vi = kw.get('don_vi', '')
            
            # Build domain based on filters
            domain = []
            if selected_term:
                domain.append(['term_months', '=', int(selected_term)])
            if from_date:
                domain.append(['create_date', '>=', from_date])
            if to_date:
                domain.append(['create_date', '<=', to_date])
            # don_vi filter not applicable for transaction data
            
            # Get all records for CSV export (no pagination)
            records = report_model.search(domain, order='create_date desc')
            
            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'STT',
                'Số Hợp đồng',
                'Số TK',
                'Số TK GDCK',
                'Khách hàng',
                'Kỳ hạn',
                'Số tiền',
                'Ngày Hợp đồng',
                'Ngày đến hạn',
                'NVCS',
                'Đơn vị'
            ])
            
            # Write data rows
            for i, record in enumerate(records, 1):
                # Sử dụng các field NAV đã tính toán sẵn từ transaction record
                nav_maturity_date = getattr(record, 'nav_maturity_date', False)
                
                # Format ngày đến hạn: ưu tiên nav_maturity_date, fallback về tính toán cũ nếu không có
                ngay_den_han = ''
                if nav_maturity_date:
                    ngay_den_han = nav_maturity_date.strftime('%d/%m/%Y')
                elif record.create_date and record.term_months:
                    # Fallback: tính từ created_at + term_months (sử dụng relativedelta logic)
                    from dateutil.relativedelta import relativedelta
                    maturity_date = record.create_date.date() + relativedelta(months=record.term_months)
                    ngay_den_han = maturity_date.strftime('%d/%m/%Y')
                
                writer.writerow([
                    i,
                    record.reference or '',
                    record.account_number or '',
                    record.account_number or '',  # Same as so_tk
                    record.investor_name or '',
                    f"{record.term_months} tháng" if record.term_months else '',
                    self._format_currency(record.amount) if record.amount else '0',
                    record.create_date.strftime('%d/%m/%Y') if record.create_date else '',
                    ngay_den_han,
                    record.approved_by.name if record.approved_by else '',
                    'HDCapital'  # Default company
                ])
            
            csv_content = output.getvalue()
            output.close()
            
            # Set response headers
            csv_http_headers = [
                ('Content-Type', 'text/csv; charset=utf-8'),
                ('Content-Length', len(csv_content.encode('utf-8'))),
                ('Content-Disposition', f'attachment; filename="Bao_cao_thong_ke_HD_ky_han_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'),
            ]
            
            return request.make_response(csv_content, headers=csv_http_headers)
            
        except Exception as e:
            print(f"Error in export_csv: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất CSV</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    def _format_currency(self, amount):
        """Helper method to format currency"""
        if not amount:
            return '0'
        return f"{amount:,.0f}".replace(',', '.')
