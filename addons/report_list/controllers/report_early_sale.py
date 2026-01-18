import json
import csv
import io
from datetime import datetime, timedelta
from odoo import http
from odoo.http import request
from odoo.addons.user_permission_management.utils.permission_checker import require_module_access

class ReportEarlySaleController(http.Controller):
    @http.route('/report-early-sale', type='http', auth='user', website=True)
    @require_module_access('report_list')
    def report_early_sale_page(self, **kw):
        """
        Route to render the Report Early Sale page.
        """
        return request.render('report_list.report_early_sale_page_template', {})

    @http.route('/report-early-sale/data', type='json', auth='user', methods=['POST'])
    def get_report_data(self, domain=None, search_values=None, limit=10, offset=0, **kw):
        """
        JSON RPC endpoint to fetch early sale data from transaction_list.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Start with the base domain from filters
            if not domain:
                domain = []

            # Add search values from inline search fields
            if search_values:
                for field, value in search_values.items():
                    if value: # Only add to domain if there is a value
                        # Map frontend field names to actual model fields
                        field_mapping = {
                            'so_hop_dong': 'reference',
                            'so_tk': 'account_number',
                            'khach_hang': 'investor_name',
                        }
                        actual_field = field_mapping.get(field, field)
                        domain.append((actual_field, 'ilike', value))

            # Resolve pagination params (page -> offset)
            page = int(kw.get('page', 1) or 1)
            limit = int(limit or 10)
            offset = int(kw.get('offset', (page - 1) * limit))

            # Build domain from filters (fund, term, date ranges for create/approved)
            filters = kw.get('filters') or {}
            fund_filter = filters.get('fund')
            term = filters.get('term')
            date_from = filters.get('dateFrom') or filters.get('from_date')
            date_to = filters.get('dateTo') or filters.get('to_date')
            sale_from = filters.get('sale_from_date')
            sale_to = filters.get('sale_to_date')

            if fund_filter:
                try:
                    fund_id_int = int(fund_filter)
                except Exception:
                    fund_id_int = fund_filter
                domain.append(('fund_id', '=', fund_id_int))
            if term:
                try:
                    domain.append(('term_months', '=', int(term)))
                except Exception:
                    pass
            if date_from:
                domain.append(('create_date', '>=', f"{date_from} 00:00:00"))
            if date_to:
                domain.append(('create_date', '<=', f"{date_to} 23:59:59"))
            if sale_from:
                domain.append(('approved_at', '>=', f"{sale_from} 00:00:00"))
            if sale_to:
                domain.append(('approved_at', '<=', f"{sale_to} 23:59:59"))

            # Debug: Log domain and parameters
            print(f"Report Early Sale Domain: {domain}")
            print(f"Limit: {limit}, Offset: {offset}, Page: {page}")
            
            # Get total count
            total = report_model.search_count(domain)
            
            # Get records with pagination
            records = report_model.search(domain, limit=limit, offset=offset, order='create_date desc')
            
            # Helper: round to nearest 50
            def _mround50(val):
                try:
                    v = float(val or 0)
                    return int(round(v / 50.0) * 50)
                except Exception:
                    return int(val or 0)

            # Format data for frontend
            data = []
            for i, record in enumerate(records, 1):
                # Calculate early sale metrics
                so_ngay_duy_tri = 0
                so_ngay_ban_truoc_han = 0
                lai_suat_truoc_han = record.interest_rate or 0
                tien_lai = 0
                lai_goc = record.amount or 0
                
                if record.create_date and record.approved_at:
                    so_ngay_duy_tri = (record.approved_at.date() - record.create_date.date()).days
                
                if record.term_months:
                    ngay_dao_han = record.create_date + timedelta(days=record.term_months * 30)
                    if record.approved_at:
                        so_ngay_ban_truoc_han = (ngay_dao_han.date() - record.approved_at.date()).days
                        # Calculate penalty
                        penalty_rate = (so_ngay_ban_truoc_han / (record.term_months * 30)) * 0.5
                        lai_suat_truoc_han = max(0, (record.interest_rate or 0) - penalty_rate)
                        
                        # Calculate interest
                        if record.amount and lai_suat_truoc_han and so_ngay_duy_tri:
                            daily_rate = lai_suat_truoc_han / 365 / 100
                            tien_lai = record.amount * daily_rate * so_ngay_duy_tri
                            lai_goc = record.amount + tien_lai
                
                data.append({
                    'id': record.id,
                    'stt': i,
                    'so_hop_dong': record.reference or '',
                    'so_tk': record.account_number or '',
                    'so_tk_gdck': record.account_number or '',
                    'khach_hang': record.investor_name or '',
                    'so_tien': _mround50(record.amount or 0),
                    'ngay_hd_mua': record.create_date.strftime('%d/%m/%Y') if record.create_date else '',
                    'ky_han': record.term_months or 0,
                    'lai_suat': record.interest_rate or 0,
                    'ngay_ban_lai_theo_hd': '',  # Not available in transaction data
                    'ngay_dao_han': (record.create_date + timedelta(days=record.term_months * 30)).strftime('%d/%m/%Y') if record.create_date and record.term_months else '',
                    'ngay_ban_lai': record.approved_at.strftime('%d/%m/%Y') if record.approved_at else '',
                    'ngay_thanh_toan': record.approved_at.strftime('%d/%m/%Y') if record.approved_at else '',
                    'so_ngay_duy_tri': so_ngay_duy_tri,
                    'so_ngay_ban_truoc_han': so_ngay_ban_truoc_han,
                    'lai_suat_truoc_han': lai_suat_truoc_han,
                    'tien_lai': _mround50(tien_lai),
                    'lai_goc': _mround50(lai_goc),
                })
            
            result = {
                'data': data,
                'total': total,
                'page': page,
                'limit': limit,
            }
            
            print(f"Returning data: {result}")
            return result
            
        except Exception as e:
            print(f"Error in get_report_data: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'error': str(e),
                'data': [],
                'total': 0,
            }

    @http.route('/report-early-sale/terms', type='json', auth='user', methods=['POST'])
    def get_terms(self, **kw):
        """
        JSON RPC endpoint to fetch terms for the filter dropdown.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Get distinct terms from transaction records
            terms = report_model.read_group([], ['term_months'], ['term_months'])
            products = []
            
            for term in terms:
                if term.get('term_months'):
                    term_value = term['term_months']
                    products.append({
                        'id': term_value, 
                        'name': f"{term_value} tháng"
                    })
            
            # Sort products by term
            products.sort(key=lambda x: x['id'])
            print(f"Terms found: {products}")
            return products
            
        except Exception as e:
            print(f"Error in get_terms: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    @http.route('/report-early-sale/export-pdf', type='http', auth='user', website=True)
    def export_pdf(self, **kw):
        """
        Route to export the early sale data as a PDF.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            selected_term = kw.get('term', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            sale_from_date = kw.get('sale_from_date', '')
            sale_to_date = kw.get('sale_to_date', '')
            
            # Build domain based on filters
            domain = []
            if selected_term:
                domain.append(['term_months', '=', int(selected_term)])
            if from_date:
                domain.append(['create_date', '>=', from_date])
            if to_date:
                domain.append(['create_date', '<=', to_date])
            if sale_from_date:
                domain.append(['approved_at', '>=', sale_from_date])
            if sale_to_date:
                domain.append(['approved_at', '<=', sale_to_date])
            
            # Get all records for PDF export (no pagination)
            records = report_model.search(domain, order='approved_at desc')
            
            # Prepare data for template
            data = {
                'records': records,
                'total_records': len(records),
                'selected_term': selected_term,
                'from_date': from_date,
                'to_date': to_date,
                'sale_from_date': sale_from_date,
                'sale_to_date': sale_to_date,
                'export_time': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'company_name': request.env.company.name,
                'user_name': request.env.user.name,
            }
            
            # Generate PDF report
            pdf_content, content_type = request.env['ir.actions.report']._render_qweb_pdf(
                'report_list.report_early_sale_pdf_template',
                res_ids=[1],  # Pass a dummy ID
                data=data
            )
            
            # Set response headers
            pdf_http_headers = [
                ('Content-Type', 'application/pdf'),
                ('Content-Length', len(pdf_content)),
                ('Content-Disposition', f'attachment; filename="Bao_cao_ban_truoc_han_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'),
            ]
            
            return request.make_response(pdf_content, headers=pdf_http_headers)
            
        except Exception as e:
            print(f"Error in export_pdf: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất PDF</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-early-sale/export-xlsx', type='http', auth='user', website=True)
    def export_xlsx(self, **kw):
        """
        Route to export the report early sale data as an XLSX.
        """
        try:
            # Sử dụng portfolio.transaction
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            customer = kw.get('customer', '')
            contract = kw.get('contract', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            
            # Build domain based on filters
            domain = []
            if customer:
                domain.append(['user_id.name', 'ilike', customer])
            if contract:
                domain.append(['name', 'ilike', contract])
            if from_date:
                domain.append(['created_at', '>=', from_date])
            if to_date:
                domain.append(['created_at', '<=', to_date])
            
            # Get all records for XLSX export (no pagination)
            records = report_model.search(domain, order='created_at desc')
            
            # Create XLSX content
            import io
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Report Early Sale"
            
            # Headers
            headers = ['STT', 'Số hợp đồng', 'Số TK', 'Khách hàng', 'Số tiền', 'Phí rút sớm']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for i, record in enumerate(records, 2):
                # Tính toán phí rút sớm (giả sử 2% của số tiền)
                phi_rut_som = round((record.amount or 0) * 0.02 / 50) * 50
                
                sheet.cell(row=i, column=1, value=i-1)  # STT
                sheet.cell(row=i, column=2, value=record.name or '')
                sheet.cell(row=i, column=3, value=record.user_id.name if record.user_id else '')
                sheet.cell(row=i, column=4, value=record.user_id.name if record.user_id else '')
                sheet.cell(row=i, column=5, value=record.amount or 0)
                sheet.cell(row=i, column=6, value=phi_rut_som)
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to buffer
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            # Set response headers
            xlsx_headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="report_early_sale_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'),
            ]
            
            return request.make_response(output.read(), headers=xlsx_headers)
            
        except Exception as e:
            print(f"Error in export_xlsx: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất XLSX</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-early-sale/export-csv', type='http', auth='user', website=True)
    def export_csv(self, **kw):
        """
        Route to export the early sale data as a CSV.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            selected_term = kw.get('term', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            sale_from_date = kw.get('sale_from_date', '')
            sale_to_date = kw.get('sale_to_date', '')
            
            # Build domain based on filters
            domain = []
            if selected_term:
                domain.append(['term_months', '=', int(selected_term)])
            if from_date:
                domain.append(['create_date', '>=', from_date])
            if to_date:
                domain.append(['create_date', '<=', to_date])
            if sale_from_date:
                domain.append(['approved_at', '>=', sale_from_date])
            if sale_to_date:
                domain.append(['approved_at', '<=', sale_to_date])
            
            # Get all records for CSV export (no pagination)
            records = report_model.search(domain, order='approved_at desc')
            
            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'STT',
                'Số Hợp đồng',
                'Số TK',
                'Số TK GDCK',
                'Khách hàng',
                'Số tiền',
                'Ngày HĐ mua',
                'Kỳ hạn',
                'Lãi suất',
                'Ngày bán lại theo HĐ',
                'Ngày đáo hạn',
                'Ngày bán lại',
                'Ngày thanh toán',
                'Số ngày duy trì',
                'Số ngày bán trước hạn',
                'Lãi suất trước hạn',
                'Tiền lãi',
                'Lãi + gốc'
            ])
            
            # Write data rows
            for i, record in enumerate(records, 1):
                # Calculate early sale metrics
                so_ngay_duy_tri = 0
                so_ngay_ban_truoc_han = 0
                lai_suat_truoc_han = record.interest_rate or 0
                tien_lai = 0
                lai_goc = record.amount or 0
                
                if record.create_date and record.approved_at:
                    so_ngay_duy_tri = (record.approved_at.date() - record.create_date.date()).days
                
                if record.term_months:
                    ngay_dao_han = record.create_date + timedelta(days=record.term_months * 30)
                    if record.approved_at:
                        so_ngay_ban_truoc_han = (ngay_dao_han.date() - record.approved_at.date()).days
                        # Calculate penalty
                        penalty_rate = (so_ngay_ban_truoc_han / (record.term_months * 30)) * 0.5
                        lai_suat_truoc_han = max(0, (record.interest_rate or 0) - penalty_rate)
                        
                        # Calculate interest
                        if record.amount and lai_suat_truoc_han and so_ngay_duy_tri:
                            daily_rate = lai_suat_truoc_han / 365 / 100
                            tien_lai = record.amount * daily_rate * so_ngay_duy_tri
                            lai_goc = record.amount + tien_lai
                
                writer.writerow([
                    i,
                    record.reference or '',
                    record.account_number or '',
                    record.account_number or '',
                    record.investor_name or '',
                    self._format_currency(record.amount) if record.amount else '0',
                    record.create_date.strftime('%d/%m/%Y') if record.create_date else '',
                    f"{record.term_months} tháng" if record.term_months else '',
                    f"{record.interest_rate}%" if record.interest_rate else '0%',
                    '',  # ngay_ban_lai_theo_hd not available
                    (record.create_date + timedelta(days=record.term_months * 30)).strftime('%d/%m/%Y') if record.create_date and record.term_months else '',
                    record.approved_at.strftime('%d/%m/%Y') if record.approved_at else '',
                    record.approved_at.strftime('%d/%m/%Y') if record.approved_at else '',
                    so_ngay_duy_tri,
                    so_ngay_ban_truoc_han,
                    f"{lai_suat_truoc_han}%",
                    self._format_currency(tien_lai),
                    self._format_currency(lai_goc)
                ])
            
            csv_content = output.getvalue()
            output.close()
            
            # Set response headers
            csv_http_headers = [
                ('Content-Type', 'text/csv; charset=utf-8'),
                ('Content-Length', len(csv_content.encode('utf-8'))),
                ('Content-Disposition', f'attachment; filename="Bao_cao_ban_truoc_han_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'),
            ]
            
            return request.make_response(csv_content, headers=csv_http_headers)
            
        except Exception as e:
            print(f"Error in export_csv: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất CSV</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    def _format_currency(self, amount):
        """Helper method to format currency"""
        if not amount:
            return '0'
        return f"{amount:,.0f}".replace(',', '.')
