import json
import csv
import io
from datetime import datetime
from odoo import http
from odoo.http import request
from odoo.addons.user_permission_management.utils.permission_checker import require_module_access

class ReportOrderHistoryController(http.Controller):
    @http.route('/report-order-history', type='http', auth='user', website=True)
    @require_module_access('report_list')
    def report_order_history_page(self, **kw):
        """
        Route to render the Report Order History page.
        """
        return request.render('report_list.report_order_history_page_template', {})

    @http.route('/report-order-history/data', type='json', auth='user', methods=['POST'])
    def get_report_data(self, domain=None, search_values=None, limit=10, offset=0, **kw):
        """
        JSON RPC endpoint to fetch order history data from transaction_list.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Start with the base domain from filters
            if not domain:
                domain = []

            # Add search values from inline search fields
            if search_values:
                for field, value in search_values.items():
                    if value: # Only add to domain if there is a value
                        # Map frontend field names to actual model fields
                        field_mapping = {
                            'so_tk': 'account_number',
                            'khach_hang': 'investor_name',
                            'ma_ck': 'fund_id.ticker',
                            'lenh': 'reference',
                        }
                        actual_field = field_mapping.get(field, field)
                        domain.append((actual_field, 'ilike', value))

            # Resolve pagination params (page -> offset)
            page = int(kw.get('page', 1) or 1)
            limit = int(limit or 10)
            offset = int(kw.get('offset', (page - 1) * limit))

            # Build domain from filters (fund, dateFrom, dateTo, status)
            filters = kw.get('filters') or {}
            fund_filter = filters.get('fund')
            date_from = filters.get('dateFrom') or filters.get('from_date')
            date_to = filters.get('dateTo') or filters.get('to_date')
            status = filters.get('status')

            if fund_filter:
                try:
                    fund_id_int = int(fund_filter)
                except Exception:
                    fund_id_int = fund_filter
                domain.append(('fund_id', '=', fund_id_int))

            if status:
                status_mapping = {
                    'pending': 'pending',
                    'matched': 'completed',
                    'cancelled': 'cancelled',
                }
                mapped_status = status_mapping.get(status, status)
                domain.append(('status', '=', mapped_status))

            if date_from and date_to and date_from == date_to:
                domain.append(('create_date', '>=', f"{date_from} 00:00:00"))
                domain.append(('create_date', '<=', f"{date_to} 23:59:59"))
            else:
                if date_from:
                    domain.append(('create_date', '>=', f"{date_from} 00:00:00"))
                if date_to:
                    domain.append(('create_date', '<=', f"{date_to} 23:59:59"))

            # Debug: Log domain and parameters
            print(f"Report Order History Domain: {domain}")
            print(f"Limit: {limit}, Offset: {offset}, Page: {page}, Filters: {filters}")
            
            # Get total count
            total = report_model.search_count(domain)
            
            # Get records with pagination
            records = report_model.search(domain, limit=limit, offset=offset, order='create_date desc')
            
            # Helper: round to nearest 50
            def _mround50(val):
                try:
                    v = float(val or 0)
                    return int(round(v / 50.0) * 50)
                except Exception:
                    return int(val or 0)

            # Format data for frontend
            data = []
            for record in records:
                # Map transaction status to order status
                status_mapping = {
                    'pending': 'pending',
                    'completed': 'matched',
                    'cancelled': 'cancelled',
                }
                order_status = status_mapping.get(record.status, 'pending')
                
                # Map transaction type to order type
                order_type_mapping = {
                    'buy': 'mua',
                    'sell': 'ban',
                    'exchange': 'hoan_doi',
                }
                order_type = order_type_mapping.get(record.transaction_type, 'mua')
                
                # Tính toán KL khớp và KL chờ từ matched orders
                matched_units = 0
                
                # Lấy tổng số lượng đã khớp từ matched orders
                # Kiểm tra cả matched_order_ids (cho lệnh mua) và matched_sell_order_ids (cho lệnh bán)
                if hasattr(record, 'matched_order_ids') and record.matched_order_ids:
                    matched_units += sum(order.matched_quantity for order in record.matched_order_ids)
                
                if hasattr(record, 'matched_sell_order_ids') and record.matched_sell_order_ids:
                    matched_units += sum(order.matched_quantity for order in record.matched_sell_order_ids)
                
                # Nếu không có matched orders, sử dụng field matched_units
                if matched_units == 0 and hasattr(record, 'matched_units'):
                    matched_units = record.matched_units or 0
                
                # Tính KL chờ = KL đặt - KL khớp
                remaining_units = (record.units or 0) - matched_units
                
                # Đảm bảo KL chờ không âm
                if remaining_units < 0:
                    remaining_units = 0
                
                data.append({
                    'id': record.id,
                    'gio_dat': record.create_date.strftime('%d/%m/%Y %H:%M:%S') if record.create_date else '',
                    'trang_thai': order_status,
                    'so_tk': record.account_number or '',
                    'so_tk_gdck': record.account_number or '',
                    'khach_hang': record.investor_name or '',
                    'nvcs': record.approved_by.name if record.approved_by else '',
                    'loai_lenh': order_type,
                    'lenh': record.reference or '',
                    'ma_ck': record.fund_id.ticker if record.fund_id else '',
                    'kl_dat': record.units or 0,
                    'gia_dat': _mround50(record.price or (record.current_nav or 0)),
                    'kl_khop': matched_units,
                    'gia_khop': _mround50(record.price or (record.current_nav or 0)),
                    'kl_cho': remaining_units,
                    'gia_cho': _mround50(record.price or (record.current_nav or 0)),
                    'shl': record.reference or '',
                })
            
            result = {
                'data': data,
                'total': total,
                'page': page,
                'limit': limit,
            }
            
            print(f"Returning data: {result}")
            return result
            
        except Exception as e:
            print(f"Error in get_report_data: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'error': str(e),
                'data': [],
                'total': 0,
            }

    @http.route('/report-order-history/securities', type='json', auth='user', methods=['POST'])
    def get_securities(self, **kw):
        """
        JSON RPC endpoint to fetch securities for the filter dropdown.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Get distinct securities from transaction records
            securities = report_model.read_group([], ['fund_id'], ['fund_id'])
            products = []
            
            for security in securities:
                if security.get('fund_id'):
                    fund_name = security['fund_id'][1] if isinstance(security['fund_id'], tuple) else security['fund_id']
                    products.append({
                        'id': fund_name, 
                        'name': fund_name
                    })
            
            # Sort products by name
            products.sort(key=lambda x: x['name'])
            print(f"Securities found: {products}")
            return products
            
        except Exception as e:
            print(f"Error in get_securities: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    @http.route('/report-order-history/export-pdf', type='http', auth='user', website=True)
    def export_pdf(self, **kw):
        """
        Route to export the order history data as a PDF.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            selected_security = kw.get('security', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            status = kw.get('status', '')
            
            # Build domain based on filters
            domain = []
            if selected_security:
                domain.append(['fund_id.name', '=', selected_security])
            if from_date:
                domain.append(['create_date', '>=', from_date])
            if to_date:
                domain.append(['create_date', '<=', to_date])
            if status:
                # Map status from frontend to transaction status
                status_mapping = {
                    'pending': 'pending',
                    'matched': 'completed',
                    'cancelled': 'cancelled',
                }
                mapped_status = status_mapping.get(status, status)
                domain.append(['status', '=', mapped_status])
            
            # Get all records for PDF export (no pagination)
            records = report_model.search(domain, order='create_date desc')
            
            # Prepare data for template
            data = {
                'records': records,
                'total_records': len(records),
                'selected_security': selected_security,
                'from_date': from_date,
                'to_date': to_date,
                'status': status,
                'export_time': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'company_name': request.env.company.name,
                'user_name': request.env.user.name,
            }
            
            # Generate PDF report
            pdf_content, content_type = request.env['ir.actions.report']._render_qweb_pdf(
                'report_list.report_order_history_pdf_template',
                res_ids=[1],  # Pass a dummy ID
                data=data
            )
            
            # Set response headers
            pdf_http_headers = [
                ('Content-Type', 'application/pdf'),
                ('Content-Length', len(pdf_content)),
                ('Content-Disposition', f'attachment; filename="Bao_cao_so_lenh_lich_su_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'),
            ]
            
            return request.make_response(pdf_content, headers=pdf_http_headers)
            
        except Exception as e:
            print(f"Error in export_pdf: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất PDF</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-order-history/export-xlsx', type='http', auth='user', website=True)
    def export_xlsx(self, **kw):
        """
        Route to export the report order history data as an XLSX.
        """
        try:
            # Sử dụng portfolio.transaction trực tiếp
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            selected_product = kw.get('product', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            
            # Build domain based on filters
            domain = []
            if selected_product:
                domain.append(['fund_id.name', '=', selected_product])
            if from_date:
                domain.append(['created_at', '>=', from_date])
            if to_date:
                domain.append(['created_at', '<=', to_date])
            
            # Get all records for XLSX export (no pagination)
            records = report_model.search(domain, order='created_at desc')
            
            # Create XLSX content
            import io
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Report Order History"
            
            # Headers
            headers = ['STT', 'Số TK', 'Khách hàng', 'Mã CK', 'Lệnh', 'KL đặt', 'KL khớp', 'KL chờ', 'Giá', 'Ngày', 'Trạng thái']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for i, record in enumerate(records, 2):
                # Tính toán KL khớp và KL chờ
                matched_units = 0
                if hasattr(record, 'matched_order_ids') and record.matched_order_ids:
                    matched_units += sum(order.matched_quantity for order in record.matched_order_ids)
                if hasattr(record, 'matched_sell_order_ids') and record.matched_sell_order_ids:
                    matched_units += sum(order.matched_quantity for order in record.matched_sell_order_ids)
                if matched_units == 0 and hasattr(record, 'matched_units'):
                    matched_units = record.matched_units or 0
                
                remaining_units = (record.units or 0) - matched_units
                if remaining_units < 0:
                    remaining_units = 0
                
                sheet.cell(row=i, column=1, value=i-1)  # STT
                sheet.cell(row=i, column=2, value=record.user_id.name if record.user_id else '')
                sheet.cell(row=i, column=3, value=record.user_id.name if record.user_id else '')
                sheet.cell(row=i, column=4, value=record.fund_id.name if record.fund_id else '')
                sheet.cell(row=i, column=5, value=record.name or '')
                sheet.cell(row=i, column=6, value=record.units or 0)
                sheet.cell(row=i, column=7, value=matched_units)
                sheet.cell(row=i, column=8, value=remaining_units)
                sheet.cell(row=i, column=9, value=record.price or 0)
                sheet.cell(row=i, column=10, value=record.created_at.strftime('%d/%m/%Y') if record.created_at else '')
                sheet.cell(row=i, column=11, value='Đã khớp' if matched_units > 0 else 'Chờ khớp')
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to buffer
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            # Set response headers
            xlsx_headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="report_order_history_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'),
            ]
            
            return request.make_response(output.read(), headers=xlsx_headers)
            
        except Exception as e:
            print(f"Error in export_xlsx: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất XLSX</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-order-history/export-csv', type='http', auth='user', website=True)
    def export_csv(self, **kw):
        """
        Route to export the order history data as a CSV.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            selected_security = kw.get('security', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            status = kw.get('status', '')
            
            # Build domain based on filters
            domain = []
            if selected_security:
                domain.append(['fund_id.name', '=', selected_security])
            if from_date:
                domain.append(['create_date', '>=', from_date])
            if to_date:
                domain.append(['create_date', '<=', to_date])
            if status:
                # Map status from frontend to transaction status
                status_mapping = {
                    'pending': 'pending',
                    'matched': 'completed',
                    'cancelled': 'cancelled',
                }
                mapped_status = status_mapping.get(status, status)
                domain.append(['status', '=', mapped_status])
            
            # Get all records for CSV export (no pagination)
            records = report_model.search(domain, order='create_date desc')
            
            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'Giờ đặt',
                'Trạng thái',
                'Số TK',
                'Số TK GDCK',
                'Khách hàng',
                'NVCS',
                'Loại lệnh',
                'Lệnh',
                'Mã CK',
                'KL đặt',
                'Giá đặt',
                'KL khớp',
                'Giá khớp',
                'KL chờ',
                'Giá chờ',
                'SHL'
            ])
            
            # Write data rows
            for record in records:
                # Map transaction status to order status
                status_mapping = {
                    'pending': 'Chờ khớp',
                    'completed': 'Đã khớp',
                    'cancelled': 'Đã hủy',
                }
                status_label = status_mapping.get(record.status, record.status)
                
                # Map transaction type to order type
                order_type_mapping = {
                    'buy': 'Lệnh mua',
                    'sell': 'Lệnh bán',
                    'exchange': 'Hoán đổi',
                }
                order_type_label = order_type_mapping.get(record.transaction_type, 'Lệnh mua')
                
                # Tính toán KL khớp và KL chờ từ matched orders
                matched_units = 0
                
                # Lấy tổng số lượng đã khớp từ matched orders
                # Kiểm tra cả matched_order_ids (cho lệnh mua) và matched_sell_order_ids (cho lệnh bán)
                if hasattr(record, 'matched_order_ids') and record.matched_order_ids:
                    matched_units += sum(order.matched_quantity for order in record.matched_order_ids)
                
                if hasattr(record, 'matched_sell_order_ids') and record.matched_sell_order_ids:
                    matched_units += sum(order.matched_quantity for order in record.matched_sell_order_ids)
                
                # Nếu không có matched orders, sử dụng field matched_units
                if matched_units == 0 and hasattr(record, 'matched_units'):
                    matched_units = record.matched_units or 0
                
                # Tính KL chờ = KL đặt - KL khớp
                remaining_units = (record.units or 0) - matched_units
                
                # Đảm bảo KL chờ không âm
                if remaining_units < 0:
                    remaining_units = 0
                
                writer.writerow([
                    record.create_date.strftime('%d/%m/%Y %H:%M:%S') if record.create_date else '',
                    status_label,
                    record.account_number or '',
                    record.account_number or '',
                    record.investor_name or '',
                    record.approved_by.name if record.approved_by else '',
                    order_type_label,
                    record.reference or '',
                    record.fund_id.ticker if record.fund_id else '',
                    self._format_number(record.units) if record.units else '0',
                    self._format_currency(record.price or (record.current_nav or 0)),
                    self._format_number(matched_units),
                    self._format_currency(record.price or (record.current_nav or 0)),
                    self._format_number(remaining_units),
                    self._format_currency(record.price or (record.current_nav or 0)),
                    record.reference or ''
                ])
            
            csv_content = output.getvalue()
            output.close()
            
            # Set response headers
            csv_http_headers = [
                ('Content-Type', 'text/csv; charset=utf-8'),
                ('Content-Length', len(csv_content.encode('utf-8'))),
                ('Content-Disposition', f'attachment; filename="Bao_cao_so_lenh_lich_su_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'),
            ]
            
            return request.make_response(csv_content, headers=csv_http_headers)
            
        except Exception as e:
            print(f"Error in export_csv: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất CSV</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    def _format_currency(self, amount):
        """Helper method to format currency"""
        if not amount:
            return '0'
        return f"{amount:,.0f}".replace(',', '.')

    def _format_number(self, number):
        """Helper method to format number"""
        if not number:
            return '0'
        return f"{number:,.0f}".replace(',', '.')

    def _get_status_label(self, status):
        """Helper method to get status label"""
        status_labels = {
            'pending': 'Chờ khớp',
            'matched': 'Đã khớp',
            'cancelled': 'Đã hủy',
            'partial': 'Khớp một phần',
        }
        return status_labels.get(status, status)

    def _get_order_type_label(self, order_type):
        """Helper method to get order type label"""
        order_types = {
            'mua': 'Lệnh mua',
            'ban': 'Lệnh bán',
        }
        return order_types.get(order_type, order_type)
