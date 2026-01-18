import json
import csv
import io
from datetime import datetime, timedelta
from odoo import http
from odoo.http import request
from odoo.addons.user_permission_management.utils.permission_checker import require_module_access

class ReportPurchaseContractController(http.Controller):
    @http.route('/report-purchase-contract', type='http', auth='user', website=True)
    @require_module_access('report_list')
    def report_purchase_contract_page(self, **kw):
        """
        Route to render the Report Purchase Contract page.
        """
        return request.render('report_list.report_purchase_contract_page_template', {})

    @http.route('/report-purchase-contract/data', type='json', auth='user', methods=['POST'])
    def get_report_data(self, domain=None, filters=None, search_values=None, limit=10, offset=0, **kw):
        """
        JSON RPC endpoint to fetch report purchase contract data with advanced search.
        """
        try:
            # Sử dụng portfolio.transaction từ transaction_list
            report_model = request.env['portfolio.transaction']
            
            # Start with the base domain from filters
            if not domain:
                domain = []
            
            # Chỉ lấy các hợp đồng có transaction_type = 'buy'
            domain.append(('transaction_type', '=', 'buy'))

            # Add search values from inline search fields
            if search_values:
                for field, value in search_values.items():
                    if value: # Only add to domain if there is a value
                        # Map frontend field names to actual model fields
                        field_mapping = {
                            'so_hop_dong': 'reference',
                            'so_tk': 'account_number',
                            'khach_hang': 'investor_name',
                        }
                        actual_field = field_mapping.get(field, field)
                        domain.append((actual_field, 'ilike', value))

            # Resolve pagination params (page -> offset)
            page = int(kw.get('page', 1) or 1)
            limit = int(limit or 10)
            offset = int(kw.get('offset', (page - 1) * limit))

            # Add filters (support both old filters arg and kw['filters'])
            effective_filters = {}
            if filters:
                effective_filters.update(filters)
            if kw.get('filters'):
                effective_filters.update(kw.get('filters'))

            if effective_filters:
                fund_filter = effective_filters.get('fund')
                date_from = effective_filters.get('dateFrom') or effective_filters.get('from_date')
                date_to = effective_filters.get('dateTo') or effective_filters.get('to_date')
                customer = effective_filters.get('khach_hang')
                contract = effective_filters.get('so_hop_dong')

                if fund_filter:
                    try:
                        fund_id_int = int(fund_filter)
                    except Exception:
                        fund_id_int = fund_filter
                    domain.append(('fund_id', '=', fund_id_int))
                if customer:
                    domain.append(('investor_name', 'ilike', customer))
                if contract:
                    domain.append(('reference', 'ilike', contract))
                if date_from and date_to and date_from == date_to:
                    domain.append(('create_date', '>=', f"{date_from} 00:00:00"))
                    domain.append(('create_date', '<=', f"{date_to} 23:59:59"))
                else:
                    if date_from:
                        domain.append(('create_date', '>=', f"{date_from} 00:00:00"))
                    if date_to:
                        domain.append(('create_date', '<=', f"{date_to} 23:59:59"))

            # Debug: Log domain and parameters
            print(f"Report Purchase Contract Domain: {domain}")
            print(f"Limit: {limit}, Offset: {offset}, Page: {page}, Filters: {effective_filters}")
            
            # Get total count
            total = report_model.search_count(domain)
            
            # Get records with pagination
            records = report_model.search(domain, limit=limit, offset=offset, order='create_date desc')
            
            # Format data for frontend - tính toán với nav_management nếu có
            data = []
            for i, record in enumerate(records, 1):
                so_tien = float(getattr(record, 'amount', 0) or 0)
                ky_han = getattr(record, 'term_months', 0) or 0
                lai_suat = float(getattr(record, 'interest_rate', 0) or 0)

                # Chuẩn bị input cho calculator
                tx_dict = {
                    'amount': so_tien,
                    'trade_price': so_tien,
                    'nav_value': float(getattr(record, 'current_nav', 0) or 0),
                    'interest_rate': lai_suat,
                    'units': float(getattr(record, 'units', 0) or 0),
                    'remaining_units': float(getattr(record, 'remaining_units', 0) or 0),
                    'term_months': ky_han,
                }
                try:
                    calculator = request.env['nav.transaction.calculator']
                    metrics = calculator.compute_transaction_metrics(tx_dict) or {}
                except Exception:
                    metrics = {}

                sell_value = float(metrics.get('sell_value') or 0.0)
                if sell_value <= 0:
                    # Fallback theo công thức cũ
                    tien_lai_du_kien = (so_tien * lai_suat) / 100
                    goc_lai_du_kien = so_tien + tien_lai_du_kien
                else:
                    goc_lai_du_kien = sell_value
                    tien_lai_du_kien = max(0.0, sell_value - so_tien)
                
                # Helper: round to nearest 50
                def _mround50(val):
                    try:
                        v = float(val or 0)
                        return int(round(v / 50.0) * 50)
                    except Exception:
                        return int(val or 0)

                data.append({
                    'id': record.id,
                    'stt': i,
                    'so_hop_dong': record.reference or '',
                    'so_tk': record.account_number or '',
                    'khach_hang': record.investor_name or '',
                    'so_tien': so_tien,
                    'ngay_hd_mua': record.create_date.strftime('%d/%m/%Y') if record.create_date else '',
                    'ngay_mua': '',  # Để trống - không có trong statistics
                    'ngay_thanh_toan': '',  # Để trống - không có trong statistics
                    'ky_han': ky_han,
                    'lai_suat': lai_suat,
                    'so_ngay_tham_gia': 0,  # Để trống - không có trong statistics
                    'tien_lai_du_kien': _mround50(tien_lai_du_kien),
                    'goc_lai_du_kien': goc_lai_du_kien,
                })
            
            result = {
                'data': data,
                'total': total,
                'page': page,
                'limit': limit,
            }
            
            print(f"Returning data: {result}")
            return result
            
        except Exception as e:
            print(f"Error in get_report_data: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'error': str(e),
                'data': [],
                'total': 0,
            }

    @http.route('/report-purchase-contract/export-pdf', type='http', auth='user', website=True)
    def export_pdf(self, **kw):
        """
        Route to export the report purchase contract data as a PDF.
        """
        try:
            # Sử dụng portfolio.transaction
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            customer = kw.get('customer', '')
            contract = kw.get('contract', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            
            # Build domain based on filters
            domain = []
            if customer:
                domain.append(['user_id.name', 'ilike', customer])
            if contract:
                domain.append(['name', 'ilike', contract])
            if from_date:
                domain.append(['created_at', '>=', f"{from_date} 00:00:00"])
            if to_date:
                domain.append(['created_at', '<=', f"{to_date} 23:59:59"])
            
            # Get all records for PDF export (no pagination)
            records = report_model.search(domain, order='created_at desc')
            
            # Prepare data for template
            data = {
                'records': records,
                'total_records': len(records),
                'customer': customer,
                'contract': contract,
                'from_date': from_date,
                'to_date': to_date,
                'export_time': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'company_name': request.env.company.name,
                'user_name': request.env.user.name,
            }
            
            # Generate PDF report
            pdf_content, content_type = request.env['ir.actions.report']._render_qweb_pdf(
                'report_list.report_purchase_contract_pdf_template',
                res_ids=[1],  # Pass a dummy ID
                data=data
            )
            
            # Set response headers
            pdf_http_headers = [
                ('Content-Type', 'application/pdf'),
                ('Content-Length', len(pdf_content)),
                ('Content-Disposition', f'attachment; filename="Danh_sach_HD_mua_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'),
            ]
            
            return request.make_response(pdf_content, headers=pdf_http_headers)
            
        except Exception as e:
            print(f"Error in export_pdf: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất PDF</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-purchase-contract/export-xlsx', type='http', auth='user', website=True)
    def export_xlsx(self, **kw):
        """
        Route to export the report purchase contract data as an XLSX.
        """
        try:
            # Sử dụng portfolio.transaction
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            customer = kw.get('customer', '')
            contract = kw.get('contract', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            
            # Build domain based on filters
            domain = []
            if customer:
                domain.append(['user_id.name', 'ilike', customer])
            if contract:
                domain.append(['name', 'ilike', contract])
            if from_date:
                domain.append(['created_at', '>=', from_date])
            if to_date:
                domain.append(['created_at', '<=', to_date])
            
            # Get all records for XLSX export (no pagination)
            records = report_model.search(domain, order='created_at desc')
            
            # Create XLSX content
            import io
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Report Purchase Contract"
            
            # Headers
            headers = ['STT', 'Số hợp đồng', 'Số TK', 'Khách hàng', 'Số tiền', 'Gốc + lãi dự kiến']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for i, record in enumerate(records, 2):
                # Tính toán Gốc + lãi dự kiến (mround 50)
                goc_lai_du_kien = record.amount or 0
                if goc_lai_du_kien > 0:
                    goc_lai_du_kien = round(goc_lai_du_kien / 50) * 50
                
                sheet.cell(row=i, column=1, value=i-1)  # STT
                sheet.cell(row=i, column=2, value=record.name or '')
                sheet.cell(row=i, column=3, value=record.user_id.name if record.user_id else '')
                sheet.cell(row=i, column=4, value=record.user_id.name if record.user_id else '')
                sheet.cell(row=i, column=5, value=record.amount or 0)
                sheet.cell(row=i, column=6, value=goc_lai_du_kien)
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to buffer
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            # Set response headers
            xlsx_headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="report_purchase_contract_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'),
            ]
            
            return request.make_response(output.read(), headers=xlsx_headers)
            
        except Exception as e:
            print(f"Error in export_xlsx: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất XLSX</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-purchase-contract/export-csv', type='http', auth='user', website=True)
    def export_csv(self, **kw):
        """
        Route to export the report purchase contract data as a CSV.
        """
        try:
            # Sử dụng portfolio.transaction
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            customer = kw.get('customer', '')
            contract = kw.get('contract', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            
            # Build domain based on filters
            domain = []
            if customer:
                domain.append(['user_id.name', 'ilike', customer])
            if contract:
                domain.append(['name', 'ilike', contract])
            if from_date:
                domain.append(['created_at', '>=', from_date])
            if to_date:
                domain.append(['created_at', '<=', to_date])
            
            # Get all records for CSV export (no pagination)
            records = report_model.search(domain, order='created_at desc')
            
            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'STT',
                'Số Hợp đồng',
                'Số TK',
                'Số TK GDCK',
                'Khách hàng',
                'Số tiền',
                'Ngày HĐ mua',
                'Ngày mua',
                'Ngày thanh toán',
                'Kỳ hạn',
                'Lãi suất',
                'Số ngày tham gia',
                'Tiền lãi dự kiến khi đến hạn',
                'Gốc + lãi dự kiến khi đến hạn'
            ])
            
            # Write data rows
            for index, record in enumerate(records):
                writer.writerow([
                    index + 1,
                    record.so_hop_dong or '',
                    record.so_tk or '',
                    record.so_tk_gdck or '',
                    record.khach_hang or '',
                    self._format_currency(record.so_tien) if record.so_tien else '0',
                    record.ngay_hd_mua.strftime('%d/%m/%Y') if record.ngay_hd_mua else '',
                    record.ngay_mua.strftime('%d/%m/%Y') if record.ngay_mua else '',
                    record.ngay_thanh_toan.strftime('%d/%m/%Y') if record.ngay_thanh_toan else '',
                    record.ky_han or '',
                    f"{record.lai_suat}%" if record.lai_suat else '',
                    record.so_ngay_tham_gia or '0',
                    self._format_currency(record.tien_lai_du_kien) if record.tien_lai_du_kien else '0',
                    self._format_currency(record.goc_lai_du_kien) if record.goc_lai_du_kien else '0'
                ])
            
            csv_content = output.getvalue()
            output.close()
            
            # Set response headers
            csv_http_headers = [
                ('Content-Type', 'text/csv; charset=utf-8'),
                ('Content-Length', len(csv_content.encode('utf-8'))),
                ('Content-Disposition', f'attachment; filename="Danh_sach_HD_mua_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'),
            ]
            
            return request.make_response(csv_content, headers=csv_http_headers)
            
        except Exception as e:
            print(f"Error in export_csv: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất CSV</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    def _format_currency(self, amount):
        """Helper method to format currency"""
        if not amount:
            return '0'
        return f"{amount:,.0f}".replace(',', '.')

    def _format_number(self, number):
        """Helper method to format number"""
        if not number:
            return '0'
        return f"{number:,.0f}".replace(',', '.')
