import json
import csv
import io
import base64
import os
from datetime import datetime
from odoo import http
from odoo.http import request
from odoo.tools import html_escape
from odoo.addons.user_permission_management.utils.permission_checker import require_module_access

class ReportTransactionController(http.Controller):
    @http.route('/report-transaction', type='http', auth='user', website=True)
    @require_module_access('report_list')
    def report_transaction_page(self, **kw):
        """
        Route to render the Report Transaction page.
        """
        return request.render('report_list.report_transaction_page_template', {})

    @http.route('/report-transaction/data', type='json', auth='user', methods=['POST'])
    def get_report_data(self, domain=None, search_values=None, limit=10, offset=0, **kw):
        """
        JSON RPC endpoint to fetch report transaction data with advanced search.
        """
        try:
            # Resolve pagination params (page -> offset)
            page = int(kw.get('page', 1) or 1)
            limit = int(limit or 10)
            offset = int(kw.get('offset', (page - 1) * limit))

            # Ensure valid domain structure
            if not domain or not isinstance(domain, (list, tuple)):
                domain = []

            # Build domain from incoming filters (fund, dateFrom, dateTo)
            filters = kw.get('filters') or {}
            fund_filter = filters.get('fund')
            date_from = filters.get('dateFrom') or filters.get('from_date')
            date_to = filters.get('dateTo') or filters.get('to_date')

            if fund_filter:
                try:
                    fund_id_int = int(fund_filter)
                except Exception:
                    fund_id_int = fund_filter
                domain.append(('fund_id', '=', fund_id_int))

            # Prefer filtering on created_at; if equal day, use exact match
            if date_from and date_to and date_from == date_to:
                domain.extend([
                    ('created_at', '>=', f"{date_from} 00:00:00"),
                    ('created_at', '<=', f"{date_to} 23:59:59")
                ])
            else:
                if date_from:
                    domain.append(('created_at', '>=', f"{date_from} 00:00:00"))
                if date_to:
                    domain.append(('created_at', '<=', f"{date_to} 23:59:59"))

            # Inline search by fields (account_number, customer_name, stock_code, order_type, transaction_code)
            inline_search = kw.get('search_values') or search_values or {}
            if inline_search:
                # Map UI fields to model fields
                mapping = {
                    'account_number': 'account_number',
                    'customer_name': 'investor_name',
                    'stock_code': 'fund_id.ticker',
                    'order_type': 'transaction_type',
                    'transaction_code': 'name',
                }
                for key, value in inline_search.items():
                    if not value:
                        continue
                    field_name = mapping.get(key, key)
                    # For order_type, accept Vietnamese labels 'Mua'/'Bán'
                    if key == 'order_type':
                        normalized = str(value).strip().lower()
                        if normalized in ('mua', 'buy'):
                            domain.append(('transaction_type', '=', 'buy'))
                            continue
                        if normalized in ('ban', 'bán', 'sell'):
                            domain.append(('transaction_type', '=', 'sell'))
                            continue
                    domain.append((field_name, 'ilike', value))

            # Sử dụng portfolio.transaction trực tiếp (đã bỏ sudo)
            report_model = request.env['portfolio.transaction']
            
            # Get total count
            total = report_model.search_count(domain)
            
            # Get records with pagination
            records = report_model.search(domain, limit=limit, offset=offset, order='created_at desc, id desc')
            
            # Format data for frontend
            data = []
            for record in records:
                # Get user info
                user_name = record.user_id.name if record.user_id else ''
                partner_name = record.user_id.partner_id.name if record.user_id and record.user_id.partner_id else ''
                
                # Get fund info
                fund_name = record.fund_id.name if record.fund_id else ''
                fund_ticker = record.fund_id.ticker if record.fund_id else ''
                
                # Map transaction type
                transaction_type_map = {
                    'buy': 'Mua',
                    'sell': 'Bán'
                }
                loai_lenh = transaction_type_map.get(record.transaction_type, record.transaction_type or '')
                
                # Round price to nearest 50 (mround 50)
                def _mround50(val):
                    try:
                        v = float(val or 0)
                        return int(round(v / 50.0) * 50)
                    except Exception:
                        return int(val or 0)

                data.append({
                    'id': record.id,
                    'so_tai_khoan': partner_name,
                    'nha_dau_tu': user_name,
                    'dksh': str(record.amount) if record.amount else '',
                    'quy': fund_name,
                    'chuong_trinh': fund_name,
                    'phien_giao_dich': record.created_at.strftime('%d/%m/%Y') if record.created_at else '',
                    'ma_giao_dich': record.name or '',
                    'loai_lenh': loai_lenh,
                    'so_ccq': record.units or 0,
                    'gia_tien': _mround50(record.price),
                    'tong_so_tien': record.amount or 0,
                    'chuong_trinh_ticker': fund_ticker,
                })
            
            # print(f"Returning data: {len(data)} records, total: {total}, domain: {domain}, page: {page}, limit: {limit}")
            return {
                'data': data,
                'total': total,
                'page': page,
                'limit': limit
            }
            
        except Exception as e:
            # print(f"Error in get_report_data: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'error': 'Internal server error',
                'records': [],
                'total': 0,
            }

    @http.route('/report-transaction/products', type='json', auth='user', methods=['POST'])
    def get_products(self, **kw):
        """
        JSON RPC endpoint to fetch products (funds) for the filter dropdown.
        """
        try:
            # Sử dụng portfolio.transaction trực tiếp (đã bỏ sudo)
            report_model = request.env['portfolio.transaction']
            
            # Gọi method từ model để lấy danh sách quỹ
            products = report_model.get_products()
            return products
            
        except Exception as e:
            # print(f"Error in get_products: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    @http.route('/report-transaction/export-pdf', type='http', auth='user', website=True)
    def export_report_transaction_pdf(self, **kw):
        """
        Route to export the report transaction data as a PDF.
        """
        try:
            # Sử dụng portfolio.transaction trực tiếp (đã bỏ sudo)
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            selected_product = kw.get('product', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            
            # Build domain based on filters
            domain = []
            if selected_product:
                domain.append(['fund_id.name', '=', selected_product])
            if from_date:
                domain.append(['created_at', '>=', f"{from_date} 00:00:00"])
            if to_date:
                domain.append(['created_at', '<=', f"{to_date} 23:59:59"])
            
            # Get all records for PDF export (no pagination)
            records = report_model.search(domain, order='created_at desc')
            
            # Calculate totals
            total_records = len(records)
            
            # Prepare data for template
            data = {
                'records': records,
                'total_records': total_records,
                'selected_product': selected_product,
                'from_date': from_date,
                'to_date': to_date,
                'export_time': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'company_name': request.env.company.name,
                'user_name': request.env.user.name,
            }
            
            # Generate PDF report
            pdf_content, content_type = request.env['ir.actions.report']._render_qweb_pdf(
                'report_list.report_transaction_pdf_template',
                res_ids=[1],  # Pass a dummy ID
                data=data
            )
            
            # Set response headers
            pdf_http_headers = [
                ('Content-Type', 'application/pdf'),
                ('Content-Length', len(pdf_content)),
                ('Content-Disposition', f'attachment; filename="Bao_cao_giao_dich_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'),
            ]
            
            return request.make_response(pdf_content, headers=pdf_http_headers)
            
        except Exception as e:
            # print(f"Error in export_report_transaction_pdf: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                "<h1>Error</h1><p>An internal error occurred. Please try again later.</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-transaction/export-xlsx', type='http', auth='user', website=True)
    def export_xlsx(self, **kw):
        """
        Route to export the report transaction data as an XLSX.
        """
        try:
            # Sử dụng portfolio.transaction trực tiếp (đã bỏ sudo)
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            selected_product = kw.get('product', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            
            # Build domain based on filters
            domain = []
            if selected_product:
                domain.append(['fund_id.name', '=', selected_product])
            if from_date:
                domain.append(['created_at', '>=', f"{from_date} 00:00:00"])
            if to_date:
                domain.append(['created_at', '<=', f"{to_date} 23:59:59"])
            
            # Get all records for XLSX export (no pagination)
            records = report_model.search(domain, order='created_at desc')
            
            # Create XLSX content
            import io
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Report Transaction"
            
            # Headers
            headers = ['STT', 'Số TK', 'Khách hàng', 'Mã CK', 'Lệnh', 'KL đặt', 'KL khớp', 'KL chờ', 'Giá', 'Ngày', 'Trạng thái']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for i, record in enumerate(records, 2):
                # Tính toán KL khớp và KL chờ
                matched_units = 0
                if hasattr(record, 'matched_order_ids') and record.matched_order_ids:
                    matched_units += sum(order.matched_quantity for order in record.matched_order_ids)
                if hasattr(record, 'matched_sell_order_ids') and record.matched_sell_order_ids:
                    matched_units += sum(order.matched_quantity for order in record.matched_sell_order_ids)
                if matched_units == 0 and hasattr(record, 'matched_units'):
                    matched_units = record.matched_units or 0
                
                remaining_units = (record.units or 0) - matched_units
                if remaining_units < 0:
                    remaining_units = 0
                
                sheet.cell(row=i, column=1, value=i-1)  # STT
                sheet.cell(row=i, column=2, value=record.user_id.name if record.user_id else '')
                sheet.cell(row=i, column=3, value=record.user_id.name if record.user_id else '')
                sheet.cell(row=i, column=4, value=record.fund_id.name if record.fund_id else '')
                sheet.cell(row=i, column=5, value=record.name or '')
                sheet.cell(row=i, column=6, value=record.units or 0)
                sheet.cell(row=i, column=7, value=matched_units)
                sheet.cell(row=i, column=8, value=remaining_units)
                sheet.cell(row=i, column=9, value=record.price or 0)
                sheet.cell(row=i, column=10, value=record.created_at.strftime('%d/%m/%Y') if record.created_at else '')
                sheet.cell(row=i, column=11, value='Đã khớp' if matched_units > 0 else 'Chờ khớp')
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to buffer
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            # Set response headers
            xlsx_headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="report_transaction_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'),
            ]
            
            return request.make_response(output.read(), headers=xlsx_headers)
            
        except Exception as e:
            # print(f"Error in export_xlsx: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                "<h1>Error</h1><p>An internal error occurred. Please try again later.</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-transaction/export-csv', type='http', auth='user', website=True)
    def export_report_transaction_csv(self, **kw):
        """
        Route to export the report transaction data as a CSV.
        """
        try:
            # Sử dụng portfolio.transaction trực tiếp (đã bỏ sudo)
            report_model = request.env['portfolio.transaction']
            
            # Get filters from URL parameters
            selected_product = kw.get('product', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            
            # Build domain based on filters
            domain = []
            if selected_product:
                domain.append(['fund_id.name', '=', selected_product])
            if from_date:
                domain.append(['created_at', '>=', f"{from_date} 00:00:00"])
            if to_date:
                domain.append(['created_at', '<=', f"{to_date} 23:59:59"])
            
            # Get all records for CSV export (no pagination)
            records = report_model.search(domain, order='created_at desc')
            
            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'Phiên giao dịch',
                'Số tài khoản',
                'Nhà đầu tư',
                'ĐKSH',
                'Quỹ',
                'Chương trình',
                'Mã giao dịch',
                'Loại lệnh',
                'Số CCQ',
                'Giá tiền',
                'Tổng số tiền'
            ])
            
            # Write data rows
            for record in records:
                # Calculate price from amount and units
                price = record.amount / record.units if record.units and record.amount else 0
                
                writer.writerow([
                    self._format_date(record.created_at),
                    record.user_id.partner_id.name if record.user_id and record.user_id.partner_id else '',
                    record.user_id.name if record.user_id else '',
                    str(record.dksh) if hasattr(record, 'dksh') and record.dksh else '',
                    record.fund_id.name if record.fund_id else '',
                    record.fund_id.ticker if record.fund_id and record.fund_id.ticker else record.fund_id.name if record.fund_id else '',
                    record.name or '',
                    self._get_order_type_label(record.transaction_type),
                    self._format_currency(record.units) if record.units else '0',
                    self._format_currency(price),
                    self._format_currency(record.amount) if record.amount else '0'
                ])
            
            csv_content = output.getvalue()
            output.close()
            
            # Set response headers
            csv_http_headers = [
                ('Content-Type', 'text/csv; charset=utf-8'),
                ('Content-Length', len(csv_content.encode('utf-8'))),
                ('Content-Disposition', f'attachment; filename="Bao_cao_giao_dich_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'),
            ]
            
            return request.make_response(csv_content, headers=csv_http_headers)
            
        except Exception as e:
            # print(f"Error in export_report_transaction_csv: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                "<h1>Error</h1><p>An internal error occurred. Please try again later.</p>",
                headers=[('Content-Type', 'text/html')]
            )

    def _format_currency(self, amount):
        """Helper method to format currency"""
        if not amount:
            return '0'
        return f"{amount:,.0f}".replace(',', '.')

    def _format_date(self, date_value):
        """Helper method to format date"""
        if not date_value:
            return ''
        if isinstance(date_value, str):
            try:
                date_value = datetime.strptime(date_value, '%Y-%m-%d').date()
            except:
                return date_value
        return date_value.strftime('%d/%m/%Y')

    def _get_order_type_label(self, order_type):
        """Helper method to get order type label"""
        order_types = {
            'buy': 'Lệnh mua',
            'sell': 'Lệnh bán',
            'exchange': 'Hoán đổi',
        }
        return order_types.get(order_type, order_type)

    @http.route('/report-transaction/contract/<int:tx_id>', type='http', auth='user')
    def transaction_contract(self, tx_id, download=False, **kw):
        """Route để xem và tải hợp đồng của giao dịch"""
        try:
            # Fix IDOR: Không dùng sudo, kiểm tra quyền truy cập
            tx = request.env['portfolio.transaction'].browse(tx_id)
            if not tx.exists():
                return request.not_found()
            
            # Kiểm tra quyền sở hữu (nếu là user thường)
            # Kiểm tra quyền sở hữu (nếu là user thường)
            if request.env.user.id != tx.user_id.id and not request.env.user.has_group('base.group_system'):
                return request.not_found()

            field = tx._fields.get('contract_pdf_path')
            if not field:
                return request.not_found()

            headers = [
                ('Content-Type', 'application/pdf'),
                ('X-Content-Type-Options', 'nosniff'),
            ]

            filename = f"contract_{tx_id}.pdf"
            if str(download).lower() in ('1', 'true', 'yes'):
                headers.append(('Content-Disposition', f'attachment; filename="{filename}"'))
            else:
                headers.append(('Content-Disposition', f'inline; filename="{filename}"'))

            try:
                if field.type == 'binary':
                    if not tx.contract_pdf_path:
                        return request.not_found()
                    data = base64.b64decode(tx.contract_pdf_path)
                    return request.make_response(data, headers=headers)
                else:
                    path = tx.contract_pdf_path
                    if not path or not os.path.isfile(path):
                        return request.not_found()
                    with open(path, 'rb') as f:
                        data = f.read()
                    return request.make_response(data, headers=headers)
            except Exception:
                return request.not_found()
        except Exception as e:
            # print(f"Error viewing contract: {str(e)}")
            return request.not_found()
