# -*- coding: utf-8 -*-
import io
from datetime import datetime
from odoo import http
from odoo.http import request, content_disposition
from odoo.addons.user_permission_management.utils.permission_checker import require_module_access

try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    openpyxl = None

class UserListController(http.Controller):

    @http.route('/user_list', type='http', auth='user', website=True)
    @require_module_access('report_list')
    def user_list_page(self, **kws):
        """Render trang HTML 'vỏ' để component Owl có thể được mount vào."""
        return request.render('report_list.user_list_report_page_template', {})

    @http.route('/user_list/get_data', type='json', auth='user', methods=['POST'])
    def get_user_list_data(self, filters, page, limit, search_values=None, **kws):
        """API endpoint để Owl component gọi và lấy dữ liệu JSON đã được phân trang."""
        try:
            # Lấy dữ liệu từ res.users model - chỉ lấy internal users
            domain = [
                ('share', '=', False),  # Chỉ lấy internal users, loại bỏ portal users
                ('id', '!=', 1)  # Loại bỏ admin user (id=1)
            ]
            
            # Filter theo phòng ban
            if filters.get('department'):
                domain.append(('department_id.name', 'ilike', filters['department']))
            
            # Filter theo trạng thái
            if filters.get('status'):
                if filters['status'] == 'active':
                    domain.append(('active', '=', True))
                elif filters['status'] == 'inactive':
                    domain.append(('active', '=', False))
            
            # Search values
            if search_values:
                if search_values.get('user_login'):
                    domain.append(('login', 'ilike', search_values['user_login']))
                if search_values.get('full_name'):
                    domain.append(('name', 'ilike', search_values['full_name']))
                if search_values.get('employee_code'):
                    # Tìm employee theo name trước, sau đó filter users
                    employee_ids = request.env['hr.employee'].search([
                        ('name', 'ilike', search_values['employee_code'])
                    ]).ids
                    if employee_ids:
                        domain.append(('employee_ids', 'in', employee_ids))
                    else:
                        # Nếu không tìm thấy employee, trả về kết quả rỗng
                        domain.append(('id', '=', 0))
            
            # Lấy records với pagination
            offset = (int(page) - 1) * int(limit)
            records = request.env['res.users'].search(domain, limit=int(limit), offset=offset)
            total_records = request.env['res.users'].search_count(domain)
            
            # Chuẩn hóa dữ liệu
            data = []
            for i, record in enumerate(records):
                # Lấy thông tin employee nếu có
                employee = None
                if hasattr(record, 'employee_ids') and record.employee_ids:
                    employee = record.employee_ids[0]
                elif hasattr(record, 'employee_id') and record.employee_id:
                    employee = record.employee_id
                
                data.append({
                    'id': record.id,
                    'stt': offset + i + 1,
                    'user_login': record.login or '',
                    'full_name': record.name or '',
                    'employee_code': employee.name if employee else '',
                    'department': employee.department_id.name if employee and employee.department_id else '',
                    'status': 'active' if record.active else 'inactive',
                    'create_date': record.create_date.strftime('%d/%m/%Y') if record.create_date else '',
                    'last_login': record.login_date.strftime('%d/%m/%Y %H:%M') if record.login_date else '',
                })
            
            return {
                'data': data,
                'total': total_records,
                'page': int(page),
                'limit': int(limit)
            }
            
        except Exception as e:
            print(f"Error in get_user_list_data: {e}")
            return {
                'data': [],
                'total': 0,
                'page': int(page),
                'limit': int(limit),
                'error': str(e)
            }

    @http.route('/user_list/export_pdf', type='http', auth='user', website=True, csrf=False)
    def export_pdf(self, **kw):
        """Xuất báo cáo ra file PDF."""
        try:
            # Lấy dữ liệu thật từ res.users - chỉ lấy internal users
            domain = [
                ('share', '=', False),  # Chỉ lấy internal users, loại bỏ portal users
                ('id', '!=', 1)  # Loại bỏ admin user (id=1)
            ]
            
            # Filter theo phòng ban
            if kw.get('department'):
                domain.append(('department_id.name', 'ilike', kw['department']))
            
            # Filter theo trạng thái
            if kw.get('status'):
                if kw['status'] == 'active':
                    domain.append(('active', '=', True))
                elif kw['status'] == 'inactive':
                    domain.append(('active', '=', False))
            
            # Search values
            if kw.get('user_login'):
                domain.append(('login', 'ilike', kw['user_login']))
            if kw.get('full_name'):
                domain.append(('name', 'ilike', kw['full_name']))
            if kw.get('employee_code'):
                # Tìm employee theo name trước, sau đó filter users
                employee_ids = request.env['hr.employee'].search([
                    ('name', 'ilike', kw['employee_code'])
                ]).ids
                if employee_ids:
                    domain.append(('employee_ids', 'in', employee_ids))
                else:
                    # Nếu không tìm thấy employee, trả về kết quả rỗng
                    domain.append(('id', '=', 0))
            
            records = request.env['res.users'].search(domain)
            
            # Chuẩn hóa dữ liệu cho export
            export_data = []
            for i, record in enumerate(records):
                # Lấy thông tin employee nếu có
                employee = None
                if hasattr(record, 'employee_ids') and record.employee_ids:
                    employee = record.employee_ids[0]
                elif hasattr(record, 'employee_id') and record.employee_id:
                    employee = record.employee_id
                
                export_data.append({
                    'stt': i + 1,
                    'user_login': record.login or '',
                    'full_name': record.name or '',
                    'employee_code': employee.name if employee else '',
                    'department': employee.department_id.name if employee and employee.department_id else '',
                    'status': 'active' if record.active else 'inactive',
                    'create_date': record.create_date.strftime('%d/%m/%Y') if record.create_date else '',
                    'last_login': record.login_date.strftime('%d/%m/%Y %H:%M') if record.login_date else '',
                })
            
            records = export_data
            
        except Exception as e:
            print(f"Error in export_pdf: {e}")
            records = []
        
        pdf_data = {
            'records': records,
            'report_date': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            'company': request.env.company,
        }
        
        pdf_content, _ = request.env['ir.actions.report']._render_qweb_pdf(
            'report_list.user_list_report_pdf_template',
            res_ids=[],
            data=pdf_data
        )
        
        filename = f"DanhSachNguoiDung_{datetime.now().strftime('%Y%m%d')}.pdf"
        pdf_http_headers = [
            ('Content-Type', 'application/pdf'),
            ('Content-Length', len(pdf_content)),
            ('Content-Disposition', content_disposition(filename)),
        ]
        return request.make_response(pdf_content, headers=pdf_http_headers)

    @http.route('/user_list/export_xlsx', type='http', auth='user', website=True, csrf=False)
    def export_xlsx(self, **kw):
        """Xuất báo cáo ra file XLSX."""
        if not openpyxl:
            return request.make_response(
                "Thư viện 'openpyxl' chưa được cài đặt. Vui lòng cài đặt bằng lệnh 'pip install openpyxl'.",
                headers=[('Content-Type', 'text/plain')]
            )

        try:
            # Lấy dữ liệu thật từ res.users - chỉ lấy internal users
            domain = [
                ('share', '=', False),  # Chỉ lấy internal users, loại bỏ portal users
                ('id', '!=', 1)  # Loại bỏ admin user (id=1)
            ]
            
            # Filter theo phòng ban
            if kw.get('department'):
                domain.append(('department_id.name', 'ilike', kw['department']))
            
            # Filter theo trạng thái
            if kw.get('status'):
                if kw['status'] == 'active':
                    domain.append(('active', '=', True))
                elif kw['status'] == 'inactive':
                    domain.append(('active', '=', False))
            
            # Search values
            if kw.get('user_login'):
                domain.append(('login', 'ilike', kw['user_login']))
            if kw.get('full_name'):
                domain.append(('name', 'ilike', kw['full_name']))
            if kw.get('employee_code'):
                # Tìm employee theo name trước, sau đó filter users
                employee_ids = request.env['hr.employee'].search([
                    ('name', 'ilike', kw['employee_code'])
                ]).ids
                if employee_ids:
                    domain.append(('employee_ids', 'in', employee_ids))
                else:
                    # Nếu không tìm thấy employee, trả về kết quả rỗng
                    domain.append(('id', '=', 0))
            
            records = request.env['res.users'].search(domain)
            
            # Chuẩn hóa dữ liệu cho export
            export_data = []
            for i, record in enumerate(records):
                # Lấy thông tin employee nếu có
                employee = None
                if hasattr(record, 'employee_ids') and record.employee_ids:
                    employee = record.employee_ids[0]
                elif hasattr(record, 'employee_id') and record.employee_id:
                    employee = record.employee_id
                
                export_data.append({
                    'stt': i + 1,
                    'user_login': record.login or '',
                    'full_name': record.name or '',
                    'employee_code': employee.name if employee else '',
                    'department': employee.department_id.name if employee and employee.department_id else '',
                    'status': 'active' if record.active else 'inactive',
                    'create_date': record.create_date.strftime('%d/%m/%Y') if record.create_date else '',
                    'last_login': record.login_date.strftime('%d/%m/%Y %H:%M') if record.login_date else '',
                })
            
            records = export_data
            
        except Exception as e:
            print(f"Error in export_xlsx: {e}")
            records = []

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "DanhSachNguoiDung"

        # Tiêu đề báo cáo
        sheet.merge_cells('A1:E1')
        title_cell = sheet['A1']
        title_cell.value = "DANH SÁCH NGƯỜI DÙNG"
        title_cell.font = Font(name='Arial', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[1].height = 30

        # Header của bảng (sẽ được append vào dòng 2)
        headers = ['STT', 'User', 'Họ và tên', 'Mã nhân viên', 'Phòng ban/Bộ phận']
        sheet.append(headers) 

        # Style cho Header
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='f37731', end_color='f37731', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        # Dòng header là dòng 2
        for cell in sheet[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Thêm dữ liệu vào body
        for rec in records:
            row = [
                rec.get('stt'), 
                rec.get('user_login'), 
                rec.get('full_name'), 
                rec.get('employee_code'),
                rec.get('department')
            ]
            sheet.append(row)

        # SỬA LỖI: Tự động điều chỉnh độ rộng cột, xử lý MergedCell
        for column_cells in sheet.columns:
            # Lấy column letter từ ô header (dòng 2), vì dòng 1 là merged cell
            # column_cells là một tuple các cell trong một cột. column_cells[1] là cell ở dòng thứ 2.
            try:
                column_letter = column_cells[1].column_letter
            except IndexError:
                # Bỏ qua nếu cột trống
                continue

            max_length = 0
            for cell in column_cells:
                # Bỏ qua MergedCell khi tính toán độ dài
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value:
                        cell_text = str(cell.value)
                        if len(cell_text) > max_length:
                            max_length = len(cell_text)
                except:
                    pass
            
            # Thêm padding và set độ rộng
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Lưu vào buffer
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f"DanhSachNguoiDung_{datetime.now().strftime('%Y%m%d')}.xlsx"
        xlsx_headers = [
            ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Disposition', content_disposition(filename)),
        ]
        return request.make_response(output.read(), headers=xlsx_headers)